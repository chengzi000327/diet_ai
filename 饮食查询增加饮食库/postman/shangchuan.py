from __future__ import annotations

import json
import io
import os
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl

try:
    import requests  # type: ignore
except Exception:  # pragma: no cover
    requests = None  # type: ignore


ES_INDEX_SCHEMA_FOOD_STANDARD_1: Dict[str, Any] = {
    "settings": {
        "analysis": {
            "normalizer": {
                "lc_norm": {
                    "type": "custom",
                    "filter": [
                        "lowercase",
                        "asciifolding",
                    ],
                }
            }
        }
    },
    "mappings": {
        "properties": {
            "food_id": {"type": "keyword"},
            "food_name": {
                "type": "keyword",
                "normalizer": "lc_norm",
                "fields": {"text": {"type": "text", "analyzer": "standard"}},
            },
            "all_names": {
                "type": "keyword",
                "normalizer": "lc_norm",
                "fields": {"text": {"type": "text", "analyzer": "standard"}},
            },
            "nutrition": {
                "properties": {
                    "ash": {"type": "double"},
                    "ca": {"type": "double"},
                    "calorie": {"type": "double"},
                    "carbohydrate": {"type": "double"},
                    "carotene": {"type": "double"},
                    "cholesterol": {"type": "double"},
                    "cu": {"type": "double"},
                    "dietaryfiber": {"type": "double"},
                    "fat": {"type": "double"},
                    "fe": {"type": "double"},
                    "i": {"type": "double"},
                    "k": {"type": "double"},
                    "mg": {"type": "double"},
                    "mn": {"type": "double"},
                    "mufa": {"type": "double"},
                    "na": {"type": "double"},
                    "niacin": {"type": "double"},
                    "p": {"type": "double"},
                    "protein": {"type": "double"},
                    "pufa": {"type": "double"},
                    "riboflavin": {"type": "double"},
                    "se": {"type": "double"},
                    "sfa": {"type": "double"},
                    "thiamin": {"type": "double"},
                    "vitamin_a": {"type": "double"},
                    "vitamin_c": {"type": "double"},
                    "vitamin_e": {"type": "double"},
                    "water": {"type": "double"},
                    "zn": {"type": "double"},
                }
            },
            "units": {
                "type": "nested",
                "properties": {
                    "100_grams_per_unit": {"type": "double"},
                    "all_unit_names": {
                        "type": "keyword",
                        "normalizer": "lc_norm",
                        "fields": {"text": {"type": "text", "analyzer": "standard"}},
                    },
                },
            },
        }
    },
}


def _load_workbook_any(path: str):
    """
    支持扩展名是 .xlsx/.xlsm/... 的正常 Excel 文件；
    也支持“实际是 xlsx 但后缀写成 .csv”的情况（通过 file-like 方式打开，绕过扩展名检查）。
    """
    p = Path(path)
    if p.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        # 注意：有些导出的 xlsx 在 read_only 模式下维度信息不完整，会导致只能读到第 1 行；
        # 这里默认用非 read_only，确保能完整读取数据。
        return openpyxl.load_workbook(path, read_only=False, data_only=True)
    # 注意：read_only 模式下 openpyxl 需要可 seek 的底层流，且必须在迭代完之前保持打开；
    # 这里直接读入内存，避免文件句柄提前关闭导致 ValueError: seek of closed file
    data = Path(path).read_bytes()
    return openpyxl.load_workbook(io.BytesIO(data), read_only=False, data_only=True)


_DELIMS = set("，,、;；/|")
_OPEN_BRACKETS = set("（([{【")
_CLOSE_BRACKETS = set("）)]}】")


def _split_names(s: Optional[str]) -> List[str]:
    if not s:
        return []

    # 只在“括号外”的分隔符处分割，避免把「米饭（蒸，代表值）」这种别名拆碎
    text = str(s).strip()
    buf: List[str] = []
    parts: List[str] = []
    depth = 0
    for ch in text:
        if ch in _OPEN_BRACKETS:
            depth += 1
            buf.append(ch)
            continue
        if ch in _CLOSE_BRACKETS:
            if depth > 0:
                depth -= 1
            buf.append(ch)
            continue
        if ch in _DELIMS and depth == 0:
            part = "".join(buf).strip()
            if part:
                parts.append(part)
            buf = []
            continue
        buf.append(ch)
    tail = "".join(buf).strip()
    if tail:
        parts.append(tail)

    # 去重但保序
    seen = set()
    out: List[str] = []
    for p in parts:
        if p not in seen:
            out.append(p)
            seen.add(p)
    return out


def _as_number(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _version_tokens(v: Any) -> List[str]:
    """
    Excel 里的“版本”既可能是 int(260130) 也可能是 str。
    用户可能输入 20260130，也可能输入 260130。
    这里把可能的等价 token 都算出来用于匹配。
    """
    if v is None:
        return []
    s = str(v).strip()
    if not s:
        return []
    digits = "".join(ch for ch in s if ch.isdigit())
    if not digits:
        return [s]

    tokens = {digits}
    if len(digits) == 8:
        tokens.add(digits[-6:])  # 20260130 -> 260130
    if len(digits) == 6:
        # 260130 -> 20260130（假设 20xx）
        tokens.add("20" + digits)
    return sorted(tokens)


def _iter_sheet_dict_rows(ws) -> Iterable[Dict[str, Any]]:
    it = ws.iter_rows(values_only=True)
    header = next(it, None)
    if not header:
        return
    header = [str(h).strip() if h is not None else "" for h in header]
    for row in it:
        if row is None:
            continue
        d = {header[i]: row[i] for i in range(min(len(header), len(row))) if header[i]}
        # 跳过整行空
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in d.values()):
            continue
        yield d


NUTRITION_COL_MAP: Dict[str, str] = {
    "能量(kcal/100g)": "calorie",
    "蛋白质(Protein)/g": "protein",
    "脂肪(Fat)/g": "fat",
    "碳水化合物(CHO)/g": "carbohydrate",
    "水分(Water)/g": "water",
    "胆固醇(Cholesterol)/mg": "cholesterol",
    "矿物质/g": "ash",
    # 这里字段名对齐你创建 index 时的 schema（dietaryfiber，无下划线）
    "总膳食纤维(Dietary fiber)/g": "dietaryfiber",
    "胡萝卜素(Carotene)/μg": "carotene",
    "维生素A(Vitamin A)/μg": "vitamin_a",
    "α-TE/mg": "vitamin_e",
    "硫胺素(Thiamin)/mg": "thiamin",
    "核黄素(Riboflavin)/mg": "riboflavin",
    "烟酸(Niacin)/mg": "niacin",
    "维生素C(Vitamin C)/mg": "vitamin_c",
    "钙(Ca)/mg": "ca",
    "磷(P)/mg": "p",
    "钾(K)/mg": "k",
    "钠(Na)/mg": "na",
    "镁(Mg)/mg": "mg",
    "铁(Fe)/mg": "fe",
    "锌(Zn)/mg": "zn",
    "硒(Se)/μg": "se",
    "铜(Cu)/mg": "cu",
    "锰(Mn)/mg": "mn",
    "碘(I)/μg": "i",
    "饱和脂肪酸(SFA)/%": "sfa",
    "单不饱和脂肪酸(MUFA)/%": "mufa",
    "多不饱和脂肪酸(PUFA)/%": "pufa",
}


def build_es_food_docs_from_standard_food_xlsx(
    xlsx_path: str,
    *,
    nutrition_sheet: str = "食物库-标准单位（100g）",
    unit_sheet: str = "食物库-单位映射系数",
    target_version: Optional[str] = None,
    version_column: str = "版本",
    food_id_mode: str = "filtered",
    food_id_width: int = 6,
    food_id_start: int = 1,
) -> List[Dict[str, Any]]:
    """
    读取“标准食物库.xlsx(可能后缀是 .csv)”并生成可写入 ES 的文档列表。

    food_id 规则：严格按 `nutrition_sheet` 中 food_name 的出现顺序，从 food_id_start 开始递增，
    使用 food_id_width 补零（例如 000001）。
    """
    wb = _load_workbook_any(xlsx_path)
    ws_nut = wb[nutrition_sheet]
    ws_unit = wb[unit_sheet]

    # 1) 单位/别名：按 food_name 聚合
    aliases_by_food: Dict[str, List[str]] = {}
    units_by_food: Dict[str, List[Tuple[str, List[str], float]]] = defaultdict(list)
    allowed_food_names: Optional[set] = None
    target_tokens = set(_version_tokens(target_version)) if target_version else set()
    if target_version:
        allowed_food_names = set()

    for r in _iter_sheet_dict_rows(ws_unit):
        food_name = str(r.get("food_name") or "").strip()
        if not food_name:
            continue

        if target_version:
            row_tokens = set(_version_tokens(r.get(version_column)))
            if not (row_tokens & target_tokens):
                continue
            assert allowed_food_names is not None
            allowed_food_names.add(food_name)

        # 别名（同一个食物会重复出现，多行相同，这里聚合去重）
        if food_name not in aliases_by_food:
            aliases_by_food[food_name] = _split_names(r.get("别名"))
        else:
            # 追加（有些版本可能不同单元格别名不完全一致）
            for a in _split_names(r.get("别名")):
                if a not in aliases_by_food[food_name]:
                    aliases_by_food[food_name].append(a)

        unit = str(r.get("unit") or "").strip()
        if not unit:
            continue
        unit_aliases = _split_names(r.get("单位别名"))
        coef = _as_number(r.get("单位转换（100克）"))
        if coef is None:
            continue

        # 合并同一单位的多行（如果重复出现）
        merged_names = [unit] + unit_aliases
        # 去重保序
        seen_u = set()
        deduped: List[str] = []
        for x in merged_names:
            if not x or x in seen_u:
                continue
            deduped.append(x)
            seen_u.add(x)
        merged_names = deduped

        units_by_food[food_name].append((unit, merged_names, float(coef)))

    # 2) 生成文档：严格跟随 nutrition_sheet 顺序
    docs: List[Dict[str, Any]] = []
    food_id = food_id_start
    nutrition_pos = food_id_start
    seen_nutrition_foods: set[str] = set()

    for r in _iter_sheet_dict_rows(ws_nut):
        food_name = str(r.get("food_name") or "").strip()
        if not food_name:
            continue
        # 防御：如果营养表里存在重复 food_name，只按第一次出现计序
        if food_name in seen_nutrition_foods:
            continue
        seen_nutrition_foods.add(food_name)

        if food_id_mode == "absolute":
            current_id = str(nutrition_pos).zfill(food_id_width)
        else:
            # 默认：filtered —— 只对“最终输出的文档”编号
            current_id = str(food_id).zfill(food_id_width)

        if allowed_food_names is not None and food_name not in allowed_food_names:
            nutrition_pos += 1
            continue

        doc: Dict[str, Any] = {
            "food_id": current_id,
            "food_name": food_name,
            "all_names": [food_name] + aliases_by_food.get(food_name, []),
            "nutrition": {},
            "units": [],
        }

        # nutrition：如果 Excel 为空则写入 None（JSON 序列化后为 null），用于“更新覆盖”ES
        nutrition: Dict[str, Optional[float]] = {}
        for col, key in NUTRITION_COL_MAP.items():
            nutrition[key] = _as_number(r.get(col))
        doc["nutrition"] = nutrition

        # units：合并同单位；如果没有，给一个默认 g
        per_food_units = units_by_food.get(food_name, [])
        if not per_food_units:
            doc["units"] = [{"all_unit_names": ["g", "克", "G"], "100_grams_per_unit": 0.01}]
        else:
            # 用 (coef, 主单位) 来避免不同单位但同 coef 冲突
            unit_map2: Dict[Tuple[str, float], Dict[str, Any]] = {}
            for unit, all_unit_names, coef in per_food_units:
                k = (unit, float(coef))
                if k not in unit_map2:
                    unit_map2[k] = {
                        "all_unit_names": list(all_unit_names),
                        "100_grams_per_unit": float(coef),
                    }
                else:
                    for n in all_unit_names:
                        if n not in unit_map2[k]["all_unit_names"]:
                            unit_map2[k]["all_unit_names"].append(n)
            doc["units"] = list(unit_map2.values())

        docs.append(doc)
        if food_id_mode != "absolute":
            food_id += 1
        nutrition_pos += 1

    return docs


def write_es_jsonl(docs: List[Dict[str, Any]], out_path: str) -> None:
    """
    每行一个文档（jsonl），适合你自己再用脚本批量写 ES。
    """
    p = Path(out_path)
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("w", encoding="utf-8") as f:
        for d in docs:
            f.write(json.dumps(d, ensure_ascii=False) + "\n")


def write_es_bulk_ndjson(
    docs: List[Dict[str, Any]],
    *,
    out_path: str,
    index_name: str,
    id_field: str = "food_id",
) -> None:
    """
    生成 ES Bulk API 的 ndjson（两行一条：action + doc）。
    """
    p = Path(out_path)
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("w", encoding="utf-8") as f:
        for d in docs:
            _id = d.get(id_field)
            action = {"index": {"_index": index_name}}
            if _id is not None:
                action["index"]["_id"] = str(_id)
            f.write(json.dumps(action, ensure_ascii=False) + "\n")
            f.write(json.dumps(d, ensure_ascii=False) + "\n")


def es_put_doc(
    *,
    es_base_url: str,
    index_name: str,
    doc_id: str,
    doc: Dict[str, Any],
    headers: Optional[Dict[str, str]] = None,
    timeout: int = 60,
) -> Dict[str, Any]:
    """
    单条写入 ES：PUT /{index}/_doc/{id}
    """
    if requests is None:
        raise RuntimeError("未安装 requests：请先 `pip install requests`")
    url = f"{es_base_url.rstrip('/')}/{index_name}/_doc/{doc_id}"
    h = {"Content-Type": "application/json"}
    if headers:
        h.update(headers)
    resp = requests.put(url, headers=h, data=json.dumps(doc, ensure_ascii=False).encode("utf-8"), timeout=timeout)
    resp.raise_for_status()
    return resp.json()


def es_create_index(
    *,
    es_base_url: str,
    index_name: str,
    schema: Dict[str, Any],
    headers: Optional[Dict[str, str]] = None,
    timeout: int = 30,
) -> Dict[str, Any]:
    """
    创建 index（schema 即 settings+mappings）。
    - 若 index 已存在：返回 {"acknowledged": True, "already_exists": True}
    """
    if requests is None:
        raise RuntimeError("未安装 requests：请先 `pip install requests`")
    url = f"{es_base_url.rstrip('/')}/{index_name}"
    h = {"Content-Type": "application/json"}
    if headers:
        h.update(headers)
    try:
        resp = requests.put(url, headers=h, data=json.dumps(schema, ensure_ascii=False).encode("utf-8"), timeout=timeout)
    except Exception as e:
        raise RuntimeError(f"创建 index 失败（网络不可达/端口拒绝/超时）。url={url}，原始错误：{e}") from e

    # 已存在通常是 400 resource_already_exists_exception（不同版本可能略有差异）
    if resp.status_code in (400, 409):
        try:
            body = resp.json()
        except Exception:
            body = {"text": resp.text}
        err = body.get("error") if isinstance(body, dict) else None
        if isinstance(err, dict) and err.get("type") == "resource_already_exists_exception":
            return {"acknowledged": True, "already_exists": True}
        resp.raise_for_status()

    resp.raise_for_status()
    return resp.json()


def es_bulk_upload_ndjson_file(
    *,
    es_base_url: str,
    bulk_ndjson_path: str,
    headers: Optional[Dict[str, str]] = None,
    timeout: int = 120,
    refresh: bool = False,
) -> Dict[str, Any]:
    """
    Bulk 写入：POST /_bulk （bulk 文件里已包含 _index/_id）

    bulk_ndjson_path: 两行一条（action + doc）的 ndjson 文件。
    """
    if requests is None:
        raise RuntimeError("未安装 requests：请先 `pip install requests`")

    url = f"{es_base_url.rstrip('/')}/_bulk"
    h = {"Content-Type": "application/x-ndjson"}
    if headers:
        h.update(headers)

    params = {"refresh": "true"} if refresh else None
    data = Path(bulk_ndjson_path).read_bytes()
    try:
        resp = requests.post(url, headers=h, params=params, data=data, timeout=timeout)
    except Exception as e:
        # 把“连接被拒绝/超时”等网络问题包装成更短的错误提示
        raise RuntimeError(
            f"请求 ES Bulk 失败（网络不可达/端口拒绝/超时）。url={url}，原始错误：{e}"
        ) from e
    resp.raise_for_status()
    result = resp.json()

    # 如果有错误，给一个更直观的异常信息
    if isinstance(result, dict) and result.get("errors"):
        items = result.get("items") or []
        bad = []
        for it in items:
            # action key 可能是 index/create/update/delete
            action = next(iter(it.values())) if isinstance(it, dict) and it else None
            if isinstance(action, dict) and action.get("error"):
                bad.append(
                    {
                        "_index": action.get("_index"),
                        "_id": action.get("_id"),
                        "status": action.get("status"),
                        "error": action.get("error"),
                    }
                )
            if len(bad) >= 5:
                break
        raise RuntimeError(f"ES bulk 返回 errors=true，示例错误（最多5条）：{bad}")

    return result


def _extract_bulk_failures(result: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    从 ES Bulk API 返回结果中提取失败项。
    返回元素示例：
      {
        "food_name": "米饭",
        "status": 400,
        "reason": "...",
        "index": "food_standard_1"
      }
    """
    failures: List[Dict[str, Any]] = []
    items = result.get("items") or []
    for it in items:
        if not isinstance(it, dict) or not it:
            continue
        action_name = next(iter(it.keys()))
        action = it.get(action_name)  # type: ignore[assignment]
        if not isinstance(action, dict):
            continue
        err = action.get("error")
        if not err:
            continue

        # 我们 bulk 文件里用 food_name 做 _id，所以这里的 _id 就是食物名
        food_name = action.get("_id")
        status = action.get("status")
        index_name = action.get("_index")

        # error reason 结构可能比较深，这里尽量给一个可读字符串
        reason = None
        if isinstance(err, dict):
            reason = err.get("reason") or err.get("type")
            if not reason:
                reason = json.dumps(err, ensure_ascii=False)
        else:
            reason = str(err)

        failures.append(
            {
                "food_name": str(food_name) if food_name is not None else "",
                "status": status,
                "reason": reason,
                "index": index_name,
            }
        )
    return failures


def print_es_bulk_summary(
    result: Dict[str, Any],
    *,
    max_failures_to_print: int = 50,
) -> Dict[str, Any]:
    """
    对 ES Bulk API 的返回结果做汇总打印：
      本次共需导入xx条，成功xx条，失败xx条
    并打印失败明细（最多 max_failures_to_print 条）。
    """
    items = result.get("items") or []
    total_docs = len(items) if isinstance(items, list) else 0
    failures = _extract_bulk_failures(result) if isinstance(result, dict) else []
    failed = len(failures)
    success = total_docs - failed

    print(f"本次共需导入{total_docs}条，成功{success}条，失败{failed}条")

    if failed > 0:
        print("失败明细（food_name | status | reason）：")
        for f in failures[:max_failures_to_print]:
            print(f"- {f.get('food_name')} | {f.get('status')} | {f.get('reason')}")
        if failed > max_failures_to_print:
            print(f"... 还有 {failed - max_failures_to_print} 条失败未打印（可调大 max_failures_to_print）")

    return {"total": total_docs, "success": success, "failed": failed, "failures": failures}


def bulk_import_foods_from_ndjson(
    *,
    es_base_url: str,
    bulk_ndjson_path: str,
    auth_env_var: str = "ES_AUTH",
    refresh: bool = True,
    print_failures: bool = True,
    max_failures_to_print: int = 200,
) -> Dict[str, Any]:
    """
    执行 bulk 导入，并在最后打印：
      本次共需导入xx条，成功xx条，失败xx条
    如果失败>0，直接在终端输出失败清单（食物名称+原因）。

    环境变量：
      ES_AUTH="Basic xxxxx"
    """
    auth = os.getenv(auth_env_var, "").strip()
    headers = {"Authorization": auth} if auth else None

    # 先探活：避免直接 bulk 时刷很长的堆栈
    if requests is None:
        raise RuntimeError("未安装 requests：请先 `pip install requests`")
    try:
        ping = requests.get(es_base_url.rstrip("/") + "/", headers=headers, timeout=10)
        # 有些集群会 401/403，但至少说明网络通
        if ping.status_code not in (200, 401, 403):
            print(f"ES 探活返回状态码：{ping.status_code}，响应：{ping.text[:200]}")
    except Exception as e:
        raise RuntimeError(
            f"无法连接到 ES（连接被拒绝/网络不通/端口未开放）。es_base_url={es_base_url}，原始错误：{e}"
        ) from e

    result = es_bulk_upload_ndjson_file(
        es_base_url=es_base_url,
        bulk_ndjson_path=bulk_ndjson_path,
        headers=headers,
        refresh=refresh,
    )

    items = result.get("items") or []
    total_docs = len(items)

    failures = _extract_bulk_failures(result)
    failed = len(failures)
    success = total_docs - failed

    # 统一汇总打印
    if print_failures:
        print_es_bulk_summary(result, max_failures_to_print=max_failures_to_print)
    else:
        print(f"本次共需导入{total_docs}条，成功{success}条，失败{failed}条")

    return {
        "total": total_docs,
        "success": success,
        "failed": failed,
        "result": result,
    }


if __name__ == "__main__":
    # 仅做汇总打印（用于你已经拿到 bulk 返回 JSON 的场景）
    # 用法：
    #   python shangchuan.py --summarize-bulk /path/to/bulk_result.json
    #   cat bulk_result.json | python shangchuan.py --summarize-bulk
    if len(sys.argv) >= 2 and sys.argv[1] == "--summarize-bulk":
        if len(sys.argv) >= 3:
            data = Path(sys.argv[2]).read_text(encoding="utf-8")
        else:
            data = sys.stdin.read()
        print_es_bulk_summary(json.loads(data))
        raise SystemExit(0)

    # 数据源（新文件）
    xlsx_path = "/Users/chengzi/Documents/work_space/饮食查询增加饮食库/postman/标准食物库0130.xlsx"
    # 只取版本 260130
    target_version = "260130"
    docs = build_es_food_docs_from_standard_food_xlsx(
        xlsx_path,
        target_version=target_version,
        food_id_mode="absolute",  # 按“全表位置”（营养表出现顺序）编号，允许跳号
    )
    print("docs:", len(docs), "first:", docs[0] if docs else None)
    # 示例输出路径（可按需修改）
    write_es_jsonl(
        docs,
        f"/Users/chengzi/Documents/work_space/饮食查询增加饮食库/postman/标准食物库0130.{target_version}.absolute.es.jsonl",
    )
    write_es_bulk_ndjson(
        docs,
        out_path=f"/Users/chengzi/Documents/work_space/饮食查询增加饮食库/postman/标准食物库0130.{target_version}.absolute.es.bulk.ndjson",
        index_name="food_standard_1",
        id_field="food_name",  # 写入 ES 时用 food_name 作为唯一 _id
    )

    # 直接 bulk 写入 ES（把 Authorization 换成你自己的）
    # 方式：把 token 写进环境变量 ES_AUTH，例如：
    #   export ES_AUTH='Basic xxxxx'
    # 然后取消下面注释即可执行导入并输出统计/失败明细（终端打印）
    # 如果要执行导入，取消注释并保持正常缩进：
    bulk_import_foods_from_ndjson(
    es_base_url="http://115.191.14.72:9200",
    bulk_ndjson_path="/Users/chengzi/Documents/work_space/饮食查询增加饮食库/postman/标准食物库.20260130.absolute.es.bulk.ndjson",
    )

    # 如果要创建 schema（index）：
    # 需要先确保 es_base_url 在你当前网络可达（你之前的 IP:port 目前是 connection refused）
    # auth = os.getenv("ES_AUTH", "").strip()
    # es_create_index(
    #     es_base_url="http://115.191.14.72:9200",
    #     index_name="food_standard_1",
    #     schema=ES_INDEX_SCHEMA_FOOD_STANDARD_1,
    #     headers={"Authorization": auth} if auth else None,
    # )

