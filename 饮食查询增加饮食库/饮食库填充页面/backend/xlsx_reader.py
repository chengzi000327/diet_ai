from __future__ import annotations

import os
import ast
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl
import re


APP_ROOT = Path(__file__).resolve().parent.parent


def clear_caches() -> None:
    """
    写入 Excel 后必须清理缓存，否则 /api/sheet 仍会返回旧数据。
    """
    try:
        _load_workbook.cache_clear()
    except Exception:
        pass
    try:
        _sheet_cache.cache_clear()
    except Exception:
        pass
    try:
        _foods_cache.cache_clear()
    except Exception:
        pass


def _discover_default_workbook() -> Optional[str]:
    """
    在未设置 WORKBOOK_PATH 时，自动从项目目录发现默认 xlsx。
    优先级：
    1) 项目根目录下的 `标准食物库参考数据0206.xlsx`
    2) 项目根目录下任意 .xlsx（按修改时间倒序取最新）
    """
    preferred = APP_ROOT / "标准食物库参考数据0206.xlsx"
    if preferred.exists():
        return str(preferred)

    xlsx_files = sorted(
        (p for p in APP_ROOT.glob("*.xlsx") if p.is_file()),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return str(xlsx_files[0]) if xlsx_files else None


def get_workbook_path() -> str:
    """
    默认从环境变量读取：
      WORKBOOK_PATH=/path/to.xlsx
    """
    p = os.getenv("WORKBOOK_PATH", "").strip()
    if not p:
        auto = _discover_default_workbook()
        if auto:
            return auto
        raise RuntimeError(
            "未设置 WORKBOOK_PATH 环境变量（指向 xlsx 文件路径），且未在项目目录找到可用的 .xlsx（例如：标准食物库参考数据0206.xlsx）"
        )
    return p


HIDDEN_COLUMNS_BY_SHEET: Dict[str, set[str]] = {
    # 用户要求：这三列不要展示/导出
    "食物库-单位映射系数": {"food_unit_count", "该单位占该食物总条数", "该食物总计数"},
}

DEFAULT_SHEETS = [
    "食物库-单位映射系数",
    "食物库-标准单位（100g）",
    "中国营养学会参考数据",
]


def _filter_columns(sheet_name: str, columns: List[str]) -> List[str]:
    hidden = HIDDEN_COLUMNS_BY_SHEET.get(sheet_name, set())
    out = [c for c in columns if c and c not in hidden]
    # 去重但保序（防止 header 重复导致前端列重复）
    seen = set()
    dedup: List[str] = []
    for c in out:
        if c in seen:
            continue
        dedup.append(c)
        seen.add(c)
    return dedup


def _filter_sheet_names(sheetnames: List[str]) -> List[str]:
    """
    可选：通过环境变量控制显示哪些 sheet
      SHEET_WHITELIST="sheet1,sheet2"
      SHEET_BLACKLIST="sheetX,sheetY"
    """
    wl = [s.strip() for s in os.getenv("SHEET_WHITELIST", "").split(",") if s.strip()]
    bl = {s.strip() for s in os.getenv("SHEET_BLACKLIST", "").split(",") if s.strip()}
    if wl:
        return [s for s in sheetnames if s in wl and s not in bl]
    if bl:
        return [s for s in sheetnames if s not in bl]
    # 默认：只展示关键的 3 张表（用户要求）
    default_present = [s for s in DEFAULT_SHEETS if s in sheetnames]
    return default_present if default_present else sheetnames


def _allowed_header_pairs(sheet_name: str, header: Tuple[Any, ...]) -> Tuple[List[str], List[Tuple[int, str]]]:
    """
    生成“允许的列名列表”和“(原始列下标, 列名)”映射。
    重要：过滤列时必须保留原始下标，否则会出现列错位（导致单位转换/核验状态/版本等读错）。
    """
    raw = [str(h).strip() if h is not None else "" for h in header]
    hidden = HIDDEN_COLUMNS_BY_SHEET.get(sheet_name, set())

    pairs: List[Tuple[int, str]] = []
    seen = set()
    for i, k in enumerate(raw):
        if not k or k in hidden:
            continue
        if k in seen:
            continue
        pairs.append((i, k))
        seen.add(k)
    cols = [k for _, k in pairs]
    return cols, pairs


def _normalize_cell(v: Any) -> Any:
    # JSON 可序列化：datetime/date 转字符串；其余保持原样；空值保持 None
    if v is None:
        return None
    # 重要：某些列（如“单位转换（100克）”）在 Excel 中是公式（例如：=300/100）。
    # openpyxl 保存过文件后，公式缓存值可能丢失；如果用 data_only=True 会读到 None。
    # 这里对“纯数字算术公式”做一次安全求值，恢复成数值。
    if isinstance(v, str) and v.startswith("="):
        expr = v[1:].strip()
        # 仅允许数字/小数点/空白/四则运算/括号，避免执行任意表达式
        if re.fullmatch(r"[0-9\.\s\+\-\*\/\(\)]+", expr or ""):
            try:
                node = ast.parse(expr, mode="eval")

                def _ok(n: ast.AST) -> bool:
                    if isinstance(n, ast.Expression):
                        return _ok(n.body)
                    if isinstance(n, ast.Constant):
                        return isinstance(n.value, (int, float))
                    if isinstance(n, ast.UnaryOp) and isinstance(n.op, (ast.UAdd, ast.USub)):
                        return _ok(n.operand)
                    if isinstance(n, ast.BinOp) and isinstance(n.op, (ast.Add, ast.Sub, ast.Mult, ast.Div)):
                        return _ok(n.left) and _ok(n.right)
                    return False

                if _ok(node):
                    val = eval(compile(node, "<formula>", "eval"), {"__builtins__": {}}, {})
                    if isinstance(val, (int, float)):
                        return float(val)
            except Exception:
                pass
    # openpyxl 可能给出 datetime/date/time
    iso = getattr(v, "isoformat", None)
    if callable(iso):
        try:
            return v.isoformat()
        except Exception:
            pass
    return v


def _tokenize_query(q: Optional[str]) -> List[str]:
    """
    将搜索串拆成多个关键词；用于“多词 AND 匹配”。
    分隔符：空白、tab、逗号/中文逗号、顿号、分号、斜杠、竖线等。
    """
    if not q:
        return []
    s = str(q).strip().lower()
    if not s:
        return []
    parts = re.split(r"[\s，,、;；/|\\\t]+", s)
    return [p for p in (x.strip() for x in parts) if p]


@lru_cache(maxsize=4)
def _load_workbook(workbook_path: str):
    p = Path(workbook_path)
    if not p.exists():
        raise FileNotFoundError(workbook_path)
    # 非 read_only：兼容某些导出文件在 read_only 下维度不完整导致只读到第1行
    # data_only=False：确保能读到公式文本；配合 _normalize_cell() 可恢复数值（如 =300/100）
    return openpyxl.load_workbook(workbook_path, read_only=False, data_only=False)


def list_sheets(workbook_path: str) -> List[str]:
    wb = _load_workbook(workbook_path)
    return _filter_sheet_names(list(wb.sheetnames))


def iter_sheet_dict_rows(
    workbook_path: str,
    sheet_name: str,
    *,
    max_rows: Optional[int] = None,
) -> Iterable[Dict[str, Any]]:
    wb = _load_workbook(workbook_path)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"sheet not found: {sheet_name}")
    ws = wb[sheet_name]

    it = ws.iter_rows(values_only=True)
    header = next(it, None)
    if not header:
        return
    _, pairs = _allowed_header_pairs(sheet_name, header)

    n = 0
    for row_idx, row in enumerate(it, start=2):
        if row is None:
            continue
        d: Dict[str, Any] = {}
        for i, k in pairs:
            if i >= len(row):
                continue
            d[k] = _normalize_cell(row[i])
        # 稳定行标识（用于前端勾选/合并）；从 2 开始（第 1 行是 header）
        d["_row"] = row_idx
        # 跳过整行空
        if not d or all(v is None or (isinstance(v, str) and not v.strip()) for v in d.values()):
            continue
        yield d
        n += 1
        if max_rows is not None and n >= max_rows:
            break


@lru_cache(maxsize=32)
def _sheet_cache(workbook_path: str, sheet_name: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    """
    简单缓存：把整个 sheet 读成 list(dict) 以便前端搜索/分页。
    对 4~5 万行规模还能接受；后续如果更大再做更高效的索引。
    """
    wb = _load_workbook(workbook_path)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"sheet not found: {sheet_name}")
    ws = wb[sheet_name]

    it = ws.iter_rows(values_only=True)
    header = next(it, None)
    if not header:
        return ([], [])
    cols, pairs = _allowed_header_pairs(sheet_name, header)

    rows: List[Dict[str, Any]] = []
    for row_idx, row in enumerate(it, start=2):
        if row is None:
            continue
        d: Dict[str, Any] = {}
        for i, k in pairs:
            if i >= len(row):
                continue
            d[k] = _normalize_cell(row[i])
        d["_row"] = row_idx
        if not d or all(v is None or (isinstance(v, str) and not v.strip()) for v in d.values()):
            continue
        rows.append(d)
    return (cols, rows)


def query_sheet_rows(
    workbook_path: str,
    sheet_name: str,
    *,
    q: Optional[str] = None,
    mode: str = "and",
    offset: int = 0,
    limit: int = 50,
) -> Dict[str, Any]:
    cols, rows = _sheet_cache(workbook_path, sheet_name)
    tokens = _tokenize_query(q)

    if tokens:
        filtered: List[Dict[str, Any]] = []
        for r in rows:
            # 多词 AND：把一行所有字段串起来做 contains
            blob = " ".join(str(v).lower() for v in r.values() if v is not None)
            if (any(t in blob for t in tokens) if str(mode).lower() == "or" else all(t in blob for t in tokens)):
                filtered.append(r)
        rows2 = filtered
    else:
        rows2 = rows

    total = len(rows2)
    offset = max(0, offset)
    limit = max(1, min(500, limit))
    page = rows2[offset : offset + limit]

    return {
        "sheet": sheet_name,
        "columns": cols,
        "total": total,
        "offset": offset,
        "limit": limit,
        "rows": page,
    }


def _split_aliases(v: Any) -> List[str]:
    """
    后端用于聚合：尽量贴近前端的拆分逻辑（兼容 tab/中文逗号/顿号等）。
    """
    if v is None:
        return []
    s = str(v).strip()
    if not s:
        return []
    import re

    parts = re.split(r"[，,、;；/|\\\n\t]+", s)
    out: List[str] = []
    seen = set()
    for p in parts:
        t = p.strip()
        if not t:
            continue
        if t in seen:
            continue
        seen.add(t)
        out.append(t)
    return out


@lru_cache(maxsize=8)
def _foods_cache(workbook_path: str, sheet_name: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    """
    把某个 sheet 聚合成“按 food_name 一条”的 foods 列表缓存。
    主要用于：单位映射系数表 -> foods 分页展示（不再按行分页，避免一页全是某个食物的多个单位）。
    """
    cols, rows = _sheet_cache(workbook_path, sheet_name)
    _ = cols  # unused

    by_food: Dict[str, Dict[str, Any]] = {}
    food_order: List[str] = []  # 保持与原表一致的出现顺序
    for r in rows:
        fn = str(r.get("food_name") or "").strip()
        if not fn:
            continue
        item = by_food.get(fn)
        if item is None:
            item = {
                "food_name": fn,
                "aliases": [],
                "units": {},  # unit -> {unit, unit_aliases, unit_to_100g, unit_to_g}
                "核验状态": r.get("核验状态"),
                "版本": r.get("版本"),
                "备注": r.get("备注"),
                "_rows": [],  # 原始行标识（_row）
            }
            by_food[fn] = item
            food_order.append(fn)

        # aliases
        for a in _split_aliases(r.get("别名")):
            if a not in item["aliases"] and a != fn:
                item["aliases"].append(a)

        # units
        unit = str(r.get("unit") or "").strip()
        if unit:
            u = item["units"].get(unit)
            if u is None:
                u = {
                    "unit": unit,
                    "unit_aliases": [],
                    "unit_to_100g": r.get("单位转换（100克）"),
                    "unit_to_g": None,
                }
                # 推导 unit_to_g
                try:
                    v100 = float(u["unit_to_100g"]) if u["unit_to_100g"] is not None and u["unit_to_100g"] != "" else None
                except Exception:
                    v100 = None
                if v100 is not None:
                    u["unit_to_g"] = v100 * 100
                item["units"][unit] = u

            for ua in _split_aliases(r.get("单位别名")):
                if ua not in u["unit_aliases"] and ua != unit:
                    u["unit_aliases"].append(ua)

        # rows
        rr = r.get("_row")
        if rr is not None:
            # 防止同一行号被重复累计到报告里
            if rr not in item["_rows"]:
                item["_rows"].append(rr)

    food_names = food_order
    foods: List[Dict[str, Any]] = []
    for fn in food_names:
        it = by_food[fn]
        units_list = list(it["units"].values())
        # 单位按名称排序
        units_list.sort(key=lambda x: str(x.get("unit") or ""))
        foods.append(
            {
                "food_name": it["food_name"],
                "aliases": it["aliases"],
                "units": units_list,
                "核验状态": it.get("核验状态"),
                "版本": it.get("版本"),
                "备注": it.get("备注"),
                "_rows": it.get("_rows", []),
            }
        )
    return (food_names, foods)


def query_foods(
    workbook_path: str,
    sheet_name: str,
    *,
    q: Optional[str] = None,
    mode: str = "and",
    offset: int = 0,
    limit: int = 50,
) -> Dict[str, Any]:
    """
    foods 维度分页查询（按 food_name 聚合）。
    q：多词 AND 匹配（大小写不敏感），匹配范围：
      - food_name / aliases
      - units.unit / units.unit_aliases
      - 版本 / 备注
    """
    _, foods = _foods_cache(workbook_path, sheet_name)
    tokens = _tokenize_query(q)
    if tokens:
        foods2: List[Dict[str, Any]] = []
        for f in foods:
            parts: List[str] = []
            parts.append(str(f.get("food_name") or ""))
            parts.extend([str(a) for a in (f.get("aliases") or [])])
            parts.append(str(f.get("版本") or ""))
            parts.append(str(f.get("备注") or ""))
            for u in f.get("units") or []:
                parts.append(str(u.get("unit") or ""))
                parts.extend([str(a) for a in (u.get("unit_aliases") or [])])
            blob = " ".join(parts).lower()
            if (any(t in blob for t in tokens) if str(mode).lower() == "or" else all(t in blob for t in tokens)):
                foods2.append(f)
        foods3 = foods2
    else:
        foods3 = foods

    total = len(foods3)
    offset = max(0, offset)
    limit = max(1, min(200, limit))
    page = foods3[offset : offset + limit]
    return {
        "sheet": sheet_name,
        "total": total,
        "offset": offset,
        "limit": limit,
        "foods": page,
    }


