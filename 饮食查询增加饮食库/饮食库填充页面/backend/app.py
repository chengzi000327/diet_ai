from __future__ import annotations

import json
import math
import os
import re
import shutil
import unicodedata
import urllib.error
import urllib.request
from collections import Counter
from difflib import SequenceMatcher
from datetime import datetime
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from fastapi import Body, FastAPI, HTTPException, Query
from fastapi.responses import FileResponse, HTMLResponse, PlainTextResponse
from fastapi.staticfiles import StaticFiles

from .xlsx_reader import clear_caches, get_workbook_path, list_sheets, query_foods, query_sheet_rows


APP_DIR = Path(__file__).resolve().parent
FRONTEND_DIR = APP_DIR.parent / "frontend"

app = FastAPI(title="饮食库填充页面")

# 静态托管前端
if FRONTEND_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(FRONTEND_DIR)), name="static")


@app.get("/", response_class=HTMLResponse)
def index():
    index_file = FRONTEND_DIR / "index.html"
    if not index_file.exists():
        return HTMLResponse("<h3>frontend/index.html not found</h3>", status_code=500)
    return HTMLResponse(index_file.read_text(encoding="utf-8"))


@app.get("/api/health", response_class=PlainTextResponse)
def health():
    return "ok"


@app.get("/api/workbook")
def workbook_info():
    try:
        p = get_workbook_path()
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"workbook_path": p, "sheets": list_sheets(p)}


@app.get("/api/sheet")
def sheet_rows(
    sheet: str = Query(..., description="sheet 名称"),
    q: Optional[str] = Query(None, description="搜索关键词（contains）"),
    mode: str = Query("and", description="搜索模式：and(默认) / or"),
    offset: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
):
    try:
        p = get_workbook_path()
        return query_sheet_rows(p, sheet, q=q, mode=mode, offset=offset, limit=limit)
    except KeyError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/foods")
def foods(
    q: Optional[str] = Query(None, description="搜索关键词（contains，匹配 food_name/aliases）"),
    offset: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
):
    """
    foods 维度分页：按 food_name 聚合（用于“合并结果表”展示全量食物，而不是按行分页）。
    目前固定使用：食物库-单位映射系数
    """
    try:
        p = get_workbook_path()
        return query_foods(p, "食物库-单位映射系数", q=q, offset=offset, limit=limit)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/sheet/export.jsonl")
def export_sheet_jsonl(sheet: str = Query(..., description="sheet 名称")):
    """
    直接把 sheet 导出为 jsonl 并下载（临时文件方式）。
    """
    try:
        p = get_workbook_path()
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    # 用缓存查询到的 rows 导出；避免重复读 excel
    data = query_sheet_rows(p, sheet, q=None, offset=0, limit=10**9)
    rows = data.get("rows", [])

    tmp_dir = APP_DIR / ".tmp_exports"
    tmp_dir.mkdir(parents=True, exist_ok=True)
    out_path = tmp_dir / f"{sheet}.jsonl"
    with out_path.open("w", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")

    return FileResponse(
        path=str(out_path),
        filename=f"{sheet}.jsonl",
        media_type="application/jsonl",
    )


@app.post("/api/nutrition/bulk_update")
def nutrition_bulk_update(payload: Dict[str, Any] = Body(...)):
    """
    将 nutrition 写回到 `食物库-标准单位（100g）`。
    重要：以 target_food_name 为准写入（也就是前端点击的那条 id）。
    如果目标 food_name 不存在，则仅针对这个 target_food_name 自动新增一行（避免按 CNS 食物名称误新增）。

    支持两种 payload：
    1) { "target_food_name": "饺子", "row": { ...CNS_row... } }
    2) { "rows": [ { "target_food_name": "饺子", "row": {...} }, ... ] }
    - 只写入目标 sheet 里存在的列
    - 写入后清理后端缓存，保证立刻可读到最新数据
    """
    try:
        p = get_workbook_path()
        # normalize inputs
        rows_in: List[Dict[str, Any]] = []
        if isinstance(payload.get("rows"), list):
            rows_in = payload.get("rows") or []
        elif payload.get("target_food_name") and isinstance(payload.get("row"), dict):
            rows_in = [{"target_food_name": payload.get("target_food_name"), "row": payload.get("row")}]
        else:
            raise ValueError('payload 必须包含 "target_food_name"+"row" 或 "rows"')

        if not rows_in:
            return {"updated": 0, "skipped": 0}

        wb = openpyxl.load_workbook(p, read_only=False, data_only=False)
        sheet_name = "食物库-标准单位（100g）"
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"sheet not found: {sheet_name}")
        ws = wb[sheet_name]

        # header -> col_index (1-based)
        header = [c.value for c in ws[1]]
        col_map: Dict[str, int] = {}
        for i, h in enumerate(header, start=1):
            if h is None:
                continue
            k = str(h).strip()
            if k and k not in col_map:
                col_map[k] = i

        if "food_name" not in col_map:
            raise RuntimeError("目标 sheet 缺少必要列：food_name")

        # build existing index: food_name -> row_idx
        idx: Dict[str, int] = {}
        fn_col = col_map["food_name"]
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=fn_col).value
            if v is None:
                continue
            name = str(v).strip()
            if name and name not in idx:
                idx[name] = r

        skip_keys = {
            "_row",
            "raw_len",
            "detail_url",
            "category_one",
            "category_two",
            "food_id",
        }

        updated = 0
        created = 0
        skipped = 0
        for item in rows_in:
            if not isinstance(item, dict):
                skipped += 1
                continue
            target_food_name = str(item.get("target_food_name") or "").strip()
            row_obj = item.get("row")
            if not target_food_name or not isinstance(row_obj, dict):
                skipped += 1
                continue

            row_idx = idx.get(target_food_name)
            is_new = False
            if row_idx is None:
                # 仅新增 target_food_name 对应行
                row_idx = ws.max_row + 1
                ws.cell(row=row_idx, column=fn_col).value = target_food_name
                idx[target_food_name] = row_idx
                is_new = True
                created += 1

            # 写入列（只写目标存在的列）
            for k, v in row_obj.items():
                if k in skip_keys:
                    continue
                if k == "食物名称":
                    continue
                kk = str(k).strip()
                if not kk:
                    continue
                if kk not in col_map:
                    continue
                ws.cell(row=row_idx, column=col_map[kk]).value = v

            # 强制写 target_food_name，避免 row_obj 里带了别的名字
            ws.cell(row=row_idx, column=fn_col).value = target_food_name
            if not is_new:
                updated += 1

        wb.save(p)
        clear_caches()
        return {"updated": updated, "created": created, "skipped": skipped}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


def _safe_num(v: Any) -> Optional[float]:
    try:
        if v is None or v == "":
            return None
        x = float(v)
        if x != x:  # NaN
            return None
        return x
    except Exception:
        return None


MERGE_CACHE_PATH = APP_DIR / ".merge_mapping.json"
_MERGE_NOISE_WORDS = (
    "约",
    "大约",
    "左右",
    "一个",
    "一份",
    "适量",
    "少许",
    "约莫",
)
_COOK_VERBS = ("炒", "煮", "蒸", "烤", "炖", "焖", "炸", "煎", "拌", "汆", "涮")


def _norm_text(s: Any) -> str:
    t = str(s or "").replace("\u00a0", " ").replace("\u3000", " ")
    t = unicodedata.normalize("NFKC", t)
    return t.strip()


def _basic_cleanup_food_name(s: Any) -> str:
    t = _norm_text(s).lower()
    if not t:
        return ""
    for w in _MERGE_NOISE_WORDS:
        t = t.replace(w, "")
    t = re.sub(r"\s+", "", t)
    t = re.sub(r"[·•]", "", t)
    return t.strip()


def _canonical_reordered_name(s: Any) -> str:
    """
    让“番茄炒蛋 / 蛋炒番茄”归到同一形式，减少顺序噪声。
    """
    t = _basic_cleanup_food_name(s)
    if not t:
        return ""
    m = re.match(r"^(.+?)(" + "|".join(_COOK_VERBS) + r")(.+)$", t)
    if not m:
        return t
    a, verb, b = _norm_text(m.group(1)), _norm_text(m.group(2)), _norm_text(m.group(3))
    if not a or not b:
        return t
    # 仅在两侧都较短时做重排，避免破坏长短语语义
    if len(a) <= 8 and len(b) <= 8:
        x, y = sorted([a, b])
        return f"{x}{verb}{y}"
    return t


def _name_merge_key(s: Any) -> str:
    return _canonical_reordered_name(s)


def _name_fingerprint(s: Any) -> Dict[str, Any]:
    t = _name_merge_key(s)
    if not t:
        return {"core": "", "process": [], "salted": False}
    process_hits: List[str] = [v for v in _COOK_VERBS if v in t]
    salted = ("咸" in t) or ("盐" in t) or ("腌" in t)
    core = t
    for v in _COOK_VERBS:
        core = core.replace(v, "")
    core = re.sub(r"[()（）\[\]【】\-_+]", "", core)
    core = _norm_text(core)
    return {
        "core": core or t,
        "process": process_hits,
        "salted": salted,
    }


def _char_ngram_counter(s: str, n: int = 2) -> Counter:
    t = _name_merge_key(s)
    if not t:
        return Counter()
    if len(t) < n:
        return Counter({t: 1})
    return Counter(t[i : i + n] for i in range(len(t) - n + 1))


def _cosine_sim(a: Counter, b: Counter) -> float:
    if not a or not b:
        return 0.0
    inter = set(a.keys()) & set(b.keys())
    num = sum(float(a[k]) * float(b[k]) for k in inter)
    den_a = math.sqrt(sum(float(v) * float(v) for v in a.values()))
    den_b = math.sqrt(sum(float(v) * float(v) for v in b.values()))
    if den_a <= 0 or den_b <= 0:
        return 0.0
    return num / (den_a * den_b)


def _vector_cosine(a: Optional[List[float]], b: Optional[List[float]]) -> Optional[float]:
    if not a or not b:
        return None
    if len(a) != len(b):
        return None
    num = 0.0
    den_a = 0.0
    den_b = 0.0
    for i in range(len(a)):
        x = float(a[i])
        y = float(b[i])
        num += x * y
        den_a += x * x
        den_b += y * y
    if den_a <= 0 or den_b <= 0:
        return None
    return num / (math.sqrt(den_a) * math.sqrt(den_b))


def _fetch_embedding_vectors(texts: List[str]) -> Dict[str, List[float]]:
    """
    OpenAI-compatible embeddings API.
    Env:
      - EMBEDDING_API_KEY
      - EMBEDDING_API_URL (default: https://api.openai.com/v1/embeddings)
      - EMBEDDING_MODEL (default: text-embedding-3-small)
      - EMBEDDING_BATCH_SIZE (default: 96)
    """
    api_key = str(os.getenv("EMBEDDING_API_KEY", "")).strip()
    if not api_key:
        return {}
    api_url = str(os.getenv("EMBEDDING_API_URL", "https://api.openai.com/v1/embeddings")).strip()
    model = str(os.getenv("EMBEDDING_MODEL", "text-embedding-3-small")).strip()
    try:
        batch_size = max(1, min(256, int(os.getenv("EMBEDDING_BATCH_SIZE", "96"))))
    except Exception:
        batch_size = 96

    uniq: List[str] = []
    seen = set()
    for t in texts:
        s = _norm_text(t)
        if not s or s in seen:
            continue
        seen.add(s)
        uniq.append(s)
    if not uniq:
        return {}

    out: Dict[str, List[float]] = {}
    for i in range(0, len(uniq), batch_size):
        chunk = uniq[i : i + batch_size]
        body = {"model": model, "input": chunk}
        req = urllib.request.Request(
            api_url,
            data=json.dumps(body, ensure_ascii=False).encode("utf-8"),
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {api_key}",
            },
            method="POST",
        )
        try:
            with urllib.request.urlopen(req, timeout=90) as resp:
                raw = resp.read().decode("utf-8", errors="ignore")
        except urllib.error.HTTPError as e:
            detail = e.read().decode("utf-8", errors="ignore")
            raise RuntimeError(f"embedding api error: {e.code} {detail}")
        obj = json.loads(raw)
        data = obj.get("data") if isinstance(obj, dict) else None
        if not isinstance(data, list):
            raise RuntimeError("embedding api 返回格式异常：缺少 data")
        for it in data:
            if not isinstance(it, dict):
                continue
            idx = int(it.get("index") or 0)
            emb = it.get("embedding")
            if not isinstance(emb, list):
                continue
            if idx < 0 or idx >= len(chunk):
                continue
            try:
                vec = [float(x) for x in emb]
            except Exception:
                continue
            out[chunk[idx]] = vec
    return out


_LOCAL_EMBED_MODEL: Any = None
_LOCAL_EMBED_MODEL_NAME: str = ""


def _get_local_embed_model() -> Any:
    """
    本地轻量向量模型加载（sentence-transformers）。
    默认模型：BAAI/bge-small-zh-v1.5
    """
    global _LOCAL_EMBED_MODEL, _LOCAL_EMBED_MODEL_NAME
    model_name = str(os.getenv("LOCAL_EMBEDDING_MODEL", "BAAI/bge-small-zh-v1.5")).strip() or "BAAI/bge-small-zh-v1.5"
    if _LOCAL_EMBED_MODEL is not None and _LOCAL_EMBED_MODEL_NAME == model_name:
        return _LOCAL_EMBED_MODEL
    try:
        from sentence_transformers import SentenceTransformer  # type: ignore
    except Exception as e:
        raise RuntimeError(f"未安装 sentence-transformers：{e}")
    _LOCAL_EMBED_MODEL = SentenceTransformer(model_name)
    _LOCAL_EMBED_MODEL_NAME = model_name
    return _LOCAL_EMBED_MODEL


def _fetch_local_embedding_vectors(texts: List[str]) -> Dict[str, List[float]]:
    """
    本地 embedding（无需外部 API）。
    Env:
      - LOCAL_EMBEDDING_MODEL (default: BAAI/bge-small-zh-v1.5)
      - LOCAL_EMBEDDING_BATCH_SIZE (default: 64)
    """
    uniq: List[str] = []
    seen = set()
    for t in texts:
        s = _norm_text(t)
        if not s or s in seen:
            continue
        seen.add(s)
        uniq.append(s)
    if not uniq:
        return {}
    try:
        bs = max(1, min(256, int(os.getenv("LOCAL_EMBEDDING_BATCH_SIZE", "64"))))
    except Exception:
        bs = 64
    model = _get_local_embed_model()
    embs = model.encode(
        uniq,
        batch_size=bs,
        show_progress_bar=False,
        normalize_embeddings=True,
    )
    out: Dict[str, List[float]] = {}
    for i, txt in enumerate(uniq):
        vec = embs[i]
        try:
            out[txt] = [float(x) for x in vec]
        except Exception:
            continue
    return out


def _kcal_diff_pct(a: Optional[float], b: Optional[float]) -> Optional[float]:
    if a is None or b is None:
        return None
    base = max(1e-6, abs((a + b) / 2.0))
    return abs(a - b) / base


def _pair_key(a: str, b: str) -> str:
    x, y = sorted([_name_merge_key(a), _name_merge_key(b)])
    return f"{x}||{y}"


def _load_merge_cache() -> Dict[str, Any]:
    if not MERGE_CACHE_PATH.exists():
        return {"alias_to_standard": {}, "pair_decisions": {}}
    try:
        obj = json.loads(MERGE_CACHE_PATH.read_text(encoding="utf-8"))
        if not isinstance(obj, dict):
            return {"alias_to_standard": {}, "pair_decisions": {}}
        a2s = obj.get("alias_to_standard") if isinstance(obj.get("alias_to_standard"), dict) else {}
        pd = obj.get("pair_decisions") if isinstance(obj.get("pair_decisions"), dict) else {}
        return {"alias_to_standard": a2s, "pair_decisions": pd}
    except Exception:
        return {"alias_to_standard": {}, "pair_decisions": {}}


def _save_merge_cache(cache: Dict[str, Any]) -> None:
    tmp = MERGE_CACHE_PATH.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(MERGE_CACHE_PATH)


def _split_aliases(s: Any) -> List[str]:
    t = _norm_text(s)
    if not t:
        return []
    out: List[str] = []
    cur = ""
    depth = 0
    for ch in t:
        if ch in ("(", "（", "[", "【"):
            depth += 1
            cur += ch
            continue
        if ch in (")", "）", "]", "】"):
            if depth > 0:
                depth -= 1
            cur += ch
            continue
        if depth == 0 and re.match(r"[，,、;；/|\\\n\t]", ch):
            v = cur.strip()
            if v:
                out.append(v)
            cur = ""
            continue
        cur += ch
    v = cur.strip()
    if v:
        out.append(v)
    return out


def _alias_tokens(parts: List[str]) -> List[str]:
    out: List[str] = []
    for p in parts:
        p2 = _norm_text(p).lower()
        if not p2:
            continue
        out.append(p2)
        for x in re.split(r"[\s,，、;；/|()（）\[\]【】\-]+", p2):
            x = _norm_text(x)
            if x:
                out.append(x)
    # keep order + unique
    seen = set()
    uniq: List[str] = []
    for x in out:
        if x in seen:
            continue
        seen.add(x)
        uniq.append(x)
    return uniq


def _ark_request_json(user_payload: Dict[str, Any], system_prompt: str) -> Dict[str, Any]:
    api_key = str(os.getenv("ARK_API_KEY", "")).strip()
    api_url = str(os.getenv("ARK_API_URL", "https://ark.cn-beijing.volces.com/api/v3/chat/completions")).strip()
    model = str(os.getenv("ARK_MODEL", "doubao-seed-1-8-251228")).strip()
    if not api_key:
        raise RuntimeError("缺少 ARK_API_KEY 环境变量")
    body = {
        "model": model,
        "temperature": 0.2,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": json.dumps(user_payload, ensure_ascii=False)},
        ],
    }
    req = urllib.request.Request(
        api_url,
        data=json.dumps(body, ensure_ascii=False).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=90) as resp:
            raw = resp.read().decode("utf-8", errors="ignore")
    except urllib.error.HTTPError as e:
        detail = e.read().decode("utf-8", errors="ignore")
        raise RuntimeError(f"model api error: {e.code} {detail}")
    data = json.loads(raw)
    content = ((((data.get("choices") or [{}])[0].get("message") or {}).get("content")) if isinstance(data, dict) else None)
    obj = _extract_first_json(str(content or ""))
    if not obj:
        raise RuntimeError("模型返回无法解析为 JSON")
    return obj


def _extract_first_json(text: str) -> Optional[Dict[str, Any]]:
    t = (text or "").strip()
    if not t:
        return None
    try:
        obj = json.loads(t)
        if isinstance(obj, dict):
            return obj
    except Exception:
        pass
    s = t.find("{")
    e = t.rfind("}")
    if s < 0 or e < 0 or e <= s:
        return None
    try:
        obj = json.loads(t[s : e + 1])
        if isinstance(obj, dict):
            return obj
    except Exception:
        return None
    return None


@app.post("/api/units/suggest_model")
def suggest_units_by_model(payload: Dict[str, Any] = Body(...)):
    """
    使用豆包模型建议单位系数（unit_to_g）。
    输入:
      {
        "food_name": "薯条",
        "all_names": ["薯条", "炸薯条"],
        "units": [{"unit":"份","unit_aliases":["小份"]}, ...]
      }
    输出:
      {
        "model": "doubao-seed-1.8",
        "items": [
          {
            "unit":"份",
            "normalized_unit":"份",
            "is_valid_unit":true,
            "unit_to_g":250,
            "confidence":0.86,
            "reason":"...",
            "source":"model"
          }
        ]
      }
    """
    try:
        model = str(os.getenv("ARK_MODEL", "doubao-seed-1-8-251228")).strip()

        food_name = str(payload.get("food_name") or "").strip()
        all_names = payload.get("all_names") or []
        units_in = payload.get("units") or []
        if not isinstance(units_in, list) or not units_in:
            raise ValueError("units 不能为空")

        units_norm: List[Dict[str, Any]] = []
        for u in units_in:
            if not isinstance(u, dict):
                continue
            unit = str(u.get("unit") or "").strip()
            if not unit:
                continue
            aliases = u.get("unit_aliases") or []
            if not isinstance(aliases, list):
                aliases = []
            units_norm.append(
                {
                    "unit": unit,
                    "unit_aliases": [str(a).strip() for a in aliases if str(a).strip()],
                }
            )
        if not units_norm:
            raise ValueError("units 为空或格式不正确")

        prompt_obj = {
            "food_name": food_name,
            "all_names": [str(x).strip() for x in (all_names or []) if str(x).strip()],
            "units": units_norm,
            "requirements": {
                "task": "estimate grams per unit",
                "must_return_json_only": True,
                "output_schema": {
                    "items": [
                        {
                            "unit": "单位名",
                            "normalized_unit": "规范单位名",
                            "is_valid_unit": True,
                            "unit_to_g": 200,
                            "confidence": 0.86,
                            "reason": "简短理由",
                            "source": "model",
                        }
                    ]
                },
                "constraints": [
                    "normalize obvious noisy units: 100g->g, 245毫升->ml, 1L->l",
                    "fix typo when obvious: 筒->桶",
                    "if it is not a real unit token, set is_valid_unit=false",
                    "unit_to_g must be number or null",
                    "confidence must be number between 0 and 1",
                    "keep source fixed as 'model'",
                ],
                "common_defaults": {
                    "g": 1,
                    "G": 1,
                    "克": 1,
                    "kg": 1000,
                    "千克": 1000,
                    "ml": 1,
                    "mL": 1,
                    "ML": 1,
                    "Ml": 1,
                    "毫升": 1,
                    "l": 1000,
                    "L": 1000,
                    "升": 1000,
                    "份": 225,
                    "小份": 150,
                    "大份": 350,
                    "碗": 200,
                    "小碗": 100,
                    "大碗": 300,
                    "两": 50,
                    "杯": 250,
                    "盘": 300,
                    "小盘": 150,
                    "大盘": 500,
                },
            },
        }

        obj = _ark_request_json(
            prompt_obj,
            "你是营养数据助手。只能输出 JSON，不要输出 markdown，不要解释。",
        )

        items = obj.get("items") if isinstance(obj, dict) else None
        if not isinstance(items, list):
            raise RuntimeError("模型返回缺少 items")

        out: List[Dict[str, Any]] = []
        for it in items:
            if not isinstance(it, dict):
                continue
            unit = str(it.get("unit") or "").strip()
            if not unit:
                continue
            normalized_unit = str(it.get("normalized_unit") or unit).strip() or unit
            is_valid_unit = bool(it.get("is_valid_unit", True))
            n = _safe_num(it.get("unit_to_g"))
            c = _safe_num(it.get("confidence"))
            if c is None:
                c = 0.0
            c = max(0.0, min(1.0, c))
            reason = str(it.get("reason") or "").strip()
            out.append(
                {
                    "unit": unit,
                    "normalized_unit": normalized_unit,
                    "is_valid_unit": is_valid_unit,
                    "unit_to_g": n,
                    "confidence": c,
                    "reason": reason,
                    "source": "model",
                }
            )
        return {"model": model, "items": out}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


def _heuristic_pair_score(a: Dict[str, Any], b: Dict[str, Any]) -> float:
    na = _name_merge_key(a.get("food_name"))
    nb = _name_merge_key(b.get("food_name"))
    if not na or not nb:
        return 0.0
    if na == nb:
        return 1.0
    toks_a = set(a.get("_tokens") or [])
    toks_b = set(b.get("_tokens") or [])
    inter = len(toks_a & toks_b)
    union = max(1, len(toks_a | toks_b))
    j = inter / union
    r = SequenceMatcher(None, na, nb).ratio()
    contain = 1.0 if (na in nb or nb in na) else 0.0
    return max(0.62 * j + 0.28 * r + 0.10 * contain, 0.78 if contain and abs(len(na) - len(nb)) <= 4 else 0.0)


def _pick_primary_name(members: List[str], freq_map: Optional[Dict[str, int]] = None) -> str:
    """
    标准名选择策略：
    1) 优先不含空格/括号/数字的候选；
    2) 评分 = 名称短(50%) + 出现频次高(50%)；
    3) 分数并列时，长度更短优先，再按字典序。
    """
    members2 = [_norm_text(x) for x in members if _norm_text(x)]
    if not members2:
        return ""

    bad_pat = re.compile(r"\s|\d|[()（）\[\]【】]")
    good = [x for x in members2 if not bad_pat.search(x)]
    candidates = good if good else members2

    fm = freq_map or {}
    max_freq = max([int(fm.get(_name_merge_key(x), 0)) for x in candidates] + [1])

    def score(x: str) -> Tuple[float, int, str]:
        key = _name_merge_key(x)
        freq = int(fm.get(key, 0))
        # 越短越高：1/(1+len)
        lx = max(1, len(key) if key else len(x))
        len_score = 1.0 / (1.0 + float(lx))
        freq_score = float(freq) / float(max_freq)
        s = 0.5 * len_score + 0.5 * freq_score
        # 返回：主分（降序）+ 次序规则（长度升序、字典序）
        return (s, lx, x)

    ranked = sorted(candidates, key=lambda x: (-score(x)[0], score(x)[1], score(x)[2]))
    return ranked[0]


def _union_find_groups(edges: List[Tuple[int, int]], n: int) -> List[List[int]]:
    parent = list(range(n))

    def find(x: int) -> int:
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a: int, b: int) -> None:
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra

    for i, j in edges:
        union(i, j)
    groups: Dict[int, List[int]] = {}
    for i in range(n):
        r = find(i)
        groups.setdefault(r, []).append(i)
    return [g for g in groups.values() if len(g) > 1]


@lru_cache(maxsize=2)
def _get_cns_kcal_index(workbook_path: str) -> Dict[str, float]:
    out: Dict[str, float] = {}
    rows = _get_cns_rows_cached(workbook_path)
    for r in rows:
        name = _name_merge_key(r.get("食物名称"))
        if not name:
            continue
        kcal = _extract_nutrition_from_cns_row(r).get("calorie")
        if kcal is None:
            continue
        if name not in out:
            out[name] = float(kcal)
    return out


@app.post("/api/merge/suggest_or_apply")
def merge_suggest_or_apply(payload: Dict[str, Any] = Body(...)):
    """
    自动合并候选编排（规则筛选 + 模型打分 + 阈值分流）。
    返回 auto_applied_groups 与 pending_review_groups，前端据此执行或确认。
    """
    try:
        t_high = float(payload.get("t_high", 0.86))
        t_low = float(payload.get("t_low", 0.70))
        max_candidates = int(payload.get("max_candidates", 500))
        merge_by_id_only = bool(payload.get("merge_by_id_only", True))
        vector_low = float(payload.get("vector_low", 0.80))
        vector_high = float(payload.get("vector_high", 0.995))
        use_cache = bool(payload.get("use_cache", True))
        use_embedding_api = bool(payload.get("use_embedding_api", False))
        use_local_embedding = bool(payload.get("use_local_embedding", True))
        p = get_workbook_path()
        foods_in = payload.get("foods")
        if not isinstance(foods_in, list) or not foods_in:
            foods_in = query_foods(p, "食物库-单位映射系数", q=None, offset=0, limit=10**9).get("foods") or []

        cache = _load_merge_cache() if use_cache else {"alias_to_standard": {}, "pair_decisions": {}}
        alias_to_standard = cache.get("alias_to_standard") if isinstance(cache.get("alias_to_standard"), dict) else {}
        pair_decisions_cache = cache.get("pair_decisions") if isinstance(cache.get("pair_decisions"), dict) else {}
        cns_kcal = _get_cns_kcal_index(p)

        rows: List[Dict[str, Any]] = []
        for f in foods_in:
            if not isinstance(f, dict):
                continue
            name = _norm_text(f.get("food_name"))
            if not name:
                continue
            aliases = [] if merge_by_id_only else [str(x).strip() for x in (f.get("aliases") or []) if str(x).strip()]
            parts = [name] + aliases
            tokens = _alias_tokens(parts)
            merge_key = _name_merge_key(name)
            canonical = _name_merge_key(alias_to_standard.get(merge_key, merge_key))
            rows.append({
                "food_name": name,
                "aliases": aliases,
                "_rows": list(dict.fromkeys([int(x) for x in (f.get("_rows") or []) if str(x).strip().lstrip("-").isdigit()])),
                "_tokens": tokens,
                "_merge_key": merge_key,
                "_canonical": canonical,
                "_vec_local": _char_ngram_counter(" ".join([name] + aliases), n=2),
                "_embed_text": _name_merge_key(" ".join([name] + aliases)),
                "_vec_api": None,
                "_fingerprint": _name_fingerprint(name),
                "_kcal": cns_kcal.get(merge_key),
            })
        n = len(rows)
        if n < 2:
            return {"auto_applied_groups": [], "pending_review_groups": [], "candidate_pairs": 0}

        # 全局名称频次（主名 + 别名）用于标准名加权选择
        name_freq: Dict[str, int] = {}
        for r in rows:
            kk = _name_merge_key(r.get("food_name"))
            if kk:
                name_freq[kk] = int(name_freq.get(kk, 0)) + 1
            if not merge_by_id_only:
                for a in (r.get("aliases") or []):
                    ka = _name_merge_key(a)
                    if ka:
                        name_freq[ka] = int(name_freq.get(ka, 0)) + 1

        embedding_ready = False
        embedding_source = "none"
        if use_embedding_api:
            try:
                emb_map = _fetch_embedding_vectors([str(r.get("_embed_text") or "") for r in rows])
                if emb_map:
                    embedding_ready = True
                    embedding_source = "api"
                    for r in rows:
                        txt = str(r.get("_embed_text") or "")
                        if txt in emb_map:
                            r["_vec_api"] = emb_map[txt]
            except Exception:
                embedding_ready = False
        if (not embedding_ready) and use_local_embedding:
            try:
                emb_map = _fetch_local_embedding_vectors([str(r.get("_embed_text") or "") for r in rows])
                if emb_map:
                    embedding_ready = True
                    embedding_source = "local_model"
                    for r in rows:
                        txt = str(r.get("_embed_text") or "")
                        if txt in emb_map:
                            r["_vec_api"] = emb_map[txt]
            except Exception:
                embedding_ready = False

        # blocking：按 token 粗筛，避免 O(n^2)
        block: Dict[str, List[int]] = {}
        for i, r in enumerate(rows):
            keys = set()
            nm = _name_merge_key(r.get("food_name"))
            if len(nm) >= 2:
                keys.add(f"p:{nm[:2]}")
                keys.add(f"s:{nm[-2:]}")
                # 增加字符片段阻塞键，提升“词序不同”召回
                for k in range(0, max(0, len(nm) - 1), 2):
                    keys.add(f"g:{nm[k:k+2]}")
            for t in (r.get("_tokens") or [])[:8]:
                if len(t) >= 2:
                    keys.add(f"t:{t[:4]}")
            for k in keys:
                block.setdefault(k, []).append(i)

        pair_seen = set()
        cands: List[Dict[str, Any]] = []
        for ids in block.values():
            if len(ids) < 2:
                continue
            ids2 = ids[:160]  # 防止单 block 过大
            for x in range(len(ids2)):
                i = ids2[x]
                for y in range(x + 1, len(ids2)):
                    j = ids2[y]
                    a, b = (i, j) if i < j else (j, i)
                    if (a, b) in pair_seen:
                        continue
                    pair_seen.add((a, b))
                    s = _heuristic_pair_score(rows[a], rows[b])
                    v_local = _cosine_sim(rows[a]["_vec_local"], rows[b]["_vec_local"])
                    v_api = _vector_cosine(rows[a].get("_vec_api"), rows[b].get("_vec_api"))
                    v_sim = float(v_api) if v_api is not None else float(v_local)
                    kdiff = _kcal_diff_pct(rows[a].get("_kcal"), rows[b].get("_kcal"))

                    # L1: 清洗后完全一致 / cache 命中 / 向量极高相似，走快速路径
                    cached_pair = pair_decisions_cache.get(_pair_key(rows[a]["food_name"], rows[b]["food_name"]))
                    prefill = None
                    if rows[a]["_canonical"] and rows[a]["_canonical"] == rows[b]["_canonical"]:
                        prefill = {"same_food": True, "confidence": 0.995, "reason": "alias cache canonical hit", "stage": "cache"}
                    elif isinstance(cached_pair, dict):
                        prefill = {
                            "same_food": bool(cached_pair.get("same_food")),
                            "confidence": max(0.0, min(1.0, float(cached_pair.get("confidence") or 0.0))),
                            "reason": _norm_text(cached_pair.get("reason")) or "pair cache hit",
                            "stage": "cache",
                        }
                    elif v_sim >= vector_high:
                        prefill = {"same_food": True, "confidence": max(0.96, float(s)), "reason": "vector exact-ish hit", "stage": "vector"}

                    if prefill is None and s < max(0.45, t_low - 0.18):
                        continue
                    if prefill is None and v_sim < vector_low and s < max(0.70, t_low):
                        continue
                    cands.append({
                        "id": len(cands) + 1,
                        "i": a,
                        "j": b,
                        "a_name": rows[a]["food_name"],
                        "b_name": rows[b]["food_name"],
                        "a_rows": rows[a].get("_rows") or [],
                        "b_rows": rows[b].get("_rows") or [],
                        "heuristic": round(float(s), 4),
                        "vector_similarity": round(float(v_sim), 4),
                        "vector_similarity_local": round(float(v_local), 4),
                        "vector_similarity_api": None if v_api is None else round(float(v_api), 4),
                        "vector_source": embedding_source if v_api is not None else "local_char",
                        "kcal_a": rows[a].get("_kcal"),
                        "kcal_b": rows[b].get("_kcal"),
                        "kcal_diff_pct": None if kdiff is None else round(float(kdiff), 4),
                        "fingerprint_a": rows[a].get("_fingerprint"),
                        "fingerprint_b": rows[b].get("_fingerprint"),
                        "prefill": prefill,
                    })
        cands.sort(key=lambda x: x["heuristic"], reverse=True)
        cands = cands[:max_candidates]

        model_res: Dict[int, Dict[str, Any]] = {}
        if cands:
            try:
                need_llm = [x for x in cands if not isinstance(x.get("prefill"), dict)]
                prompt_obj = {
                    "task": "judge whether two food names should be merged into one nutrition entity",
                    "rules": [
                        "同义词、品牌名与通用名可合并（如 可口可乐 与 可乐）",
                        "基础食材相同且处理方式不显著改变营养可合并（如 鸡蛋 与 水煮蛋）",
                        "加工深度变化/辅料显著变化通常不合并（如 鲜葡萄 与 葡萄干、鸡蛋 与 炒鸡蛋）",
                        "尽量保守，宁可不合并也不要误合并",
                    ],
                    "few_shot": payload.get("few_shot_examples") or [
                        {"a": "番茄", "b": "西红柿", "same_food": True, "reason": "同义词"},
                        {"a": "鸡蛋", "b": "鸡蛋(煮)", "same_food": True, "reason": "食材相同，处理方式不显著改变营养"},
                        {"a": "牛奶", "b": "酸奶", "same_food": False, "reason": "发酵导致性质变化"},
                        {"a": "土豆", "b": "土豆泥", "same_food": False, "reason": "常含额外辅料，营养差异明显"},
                    ],
                    "pairs": [
                        {
                            "id": x["id"],
                            "a": x["a_name"],
                            "b": x["b_name"],
                            "heuristic": x["heuristic"],
                            "vector_similarity": x.get("vector_similarity"),
                            "kcal_a": x.get("kcal_a"),
                            "kcal_b": x.get("kcal_b"),
                            "kcal_diff_pct": x.get("kcal_diff_pct"),
                            "fingerprint_a": x.get("fingerprint_a"),
                            "fingerprint_b": x.get("fingerprint_b"),
                        }
                        for x in need_llm
                    ],
                    "output_schema": {
                        "items": [{"id": 1, "same_food": True, "confidence": 0.9, "reason": "简短理由", "fingerprint": {"core": "egg", "variety": "any"}}]
                    },
                    "decision_hint": "优先判断蛋白质/脂肪/碳水是否会显著变化；若变化小可合并",
                }
                if need_llm:
                    obj = _ark_request_json(
                        prompt_obj,
                        "你是食品标准化助手。先进行内部推理，再只输出 JSON。禁止 markdown 与额外文本。",
                    )
                    items = obj.get("items") if isinstance(obj, dict) else None
                    if isinstance(items, list):
                        for it in items:
                            if not isinstance(it, dict):
                                continue
                            pid = int(it.get("id") or 0)
                            if pid <= 0:
                                continue
                            model_res[pid] = {
                                "same_food": bool(it.get("same_food")),
                                "confidence": max(0.0, min(1.0, float(it.get("confidence") or 0.0))),
                                "reason": _norm_text(it.get("reason")),
                            }
            except Exception:
                # 模型失败时退化到 heuristic
                model_res = {}

        edges_low: List[Tuple[int, int]] = []
        pair_meta: Dict[Tuple[int, int], Dict[str, Any]] = {}
        pair_details: List[Dict[str, Any]] = []
        for c in cands:
            pre = c.get("prefill") if isinstance(c.get("prefill"), dict) else None
            mr = pre or model_res.get(c["id"])
            if isinstance(mr, dict):
                same_food = bool(mr.get("same_food"))
                conf = float(mr.get("confidence") or 0.0)
                reason = _norm_text(mr.get("reason"))
            else:
                conf = float(c["heuristic"])
                same_food = conf >= t_high
                reason = "heuristic fallback"

            # kcal 差异辅助降权（若差异明显则更保守）
            kd = c.get("kcal_diff_pct")
            if kd is not None:
                try:
                    kd2 = float(kd)
                    if kd2 > 0.35:
                        conf *= 0.85
                    elif kd2 <= 0.05:
                        conf = min(1.0, conf + 0.03)
                except Exception:
                    pass

            if (not same_food) or conf < t_low:
                decision = "reject"
            elif conf >= t_high:
                decision = "auto"
            else:
                decision = "pending"
            pair_details.append(
                {
                    "id": int(c["id"]),
                    "a": c["a_name"],
                    "b": c["b_name"],
                    "a_rows": c.get("a_rows") or [],
                    "b_rows": c.get("b_rows") or [],
                    "heuristic": float(c["heuristic"]),
                    "vector_similarity": float(c.get("vector_similarity") or 0.0),
                    "vector_similarity_local": float(c.get("vector_similarity_local") or 0.0),
                    "vector_similarity_api": c.get("vector_similarity_api"),
                    "vector_source": str(c.get("vector_source") or "local"),
                    "kcal_a": c.get("kcal_a"),
                    "kcal_b": c.get("kcal_b"),
                    "kcal_diff_pct": c.get("kcal_diff_pct"),
                    "same_food": bool(same_food),
                    "confidence": round(float(conf), 4),
                    "decision": decision,
                    "reason": reason,
                    "stage": str((pre or {}).get("stage") or ("model" if model_res.get(c["id"]) else "heuristic")),
                }
            )

            # 写入 pair 决策缓存（供下次直接命中）
            if use_cache:
                pk = _pair_key(c["a_name"], c["b_name"])
                pair_decisions_cache[pk] = {
                    "same_food": bool(same_food),
                    "confidence": round(float(conf), 4),
                    "reason": reason,
                    "updated_at": datetime.now().isoformat(),
                }
            if not same_food or conf < t_low:
                continue
            i, j = int(c["i"]), int(c["j"])
            edges_low.append((i, j))
            pair_meta[(min(i, j), max(i, j))] = {
                "confidence": conf,
                "reason": reason,
                "a": c["a_name"],
                "b": c["b_name"],
            }

        groups_idx = _union_find_groups(edges_low, n)
        auto_groups: List[Dict[str, Any]] = []
        pending_groups: List[Dict[str, Any]] = []
        for gidx in groups_idx:
            members = [rows[i]["food_name"] for i in gidx]
            confs: List[float] = []
            reasons: List[str] = []
            for x in range(len(gidx)):
                for y in range(x + 1, len(gidx)):
                    a, b = sorted((gidx[x], gidx[y]))
                    m = pair_meta.get((a, b))
                    if not m:
                        continue
                    confs.append(float(m["confidence"]))
                    if m["reason"]:
                        reasons.append(str(m["reason"]))
            score = sum(confs) / len(confs) if confs else 0.0
            out = {
                "primary": _pick_primary_name(members, name_freq),
                "members": members,
                "member_rows": [
                    {"food_name": rows[i]["food_name"], "rows": rows[i].get("_rows") or []}
                    for i in gidx
                ],
                "confidence": round(score, 4),
                "reason": "；".join(reasons[:3]),
            }
            if score >= t_high:
                auto_groups.append(out)
            else:
                pending_groups.append(out)

            # 高置信组写入 alias -> standard 缓存
            if use_cache and score >= t_high:
                standard = _name_merge_key(out["primary"])
                for m in members:
                    mk = _name_merge_key(m)
                    if mk and mk != standard:
                        alias_to_standard[mk] = standard

        auto_groups.sort(key=lambda x: (-float(x["confidence"]), x["primary"]))
        pending_groups.sort(key=lambda x: (-float(x["confidence"]), x["primary"]))
        if use_cache:
            cache["alias_to_standard"] = alias_to_standard
            cache["pair_decisions"] = pair_decisions_cache
            _save_merge_cache(cache)
        return {
            "auto_applied_groups": auto_groups,
            "pending_review_groups": pending_groups,
            "candidate_pairs": len(cands),
            "candidate_pairs_detail": pair_details,
            "embedding_api_used": embedding_source == "api",
            "local_embedding_used": embedding_source == "local_model",
            "embedding_source": embedding_source if embedding_ready else "local_char",
            "thresholds": {"t_high": t_high, "t_low": t_low},
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


NUT_KEYS = [
    "calorie",
    "protein",
    "fat",
    "carbohydrate",
    "water",
    "cholesterol",
    "ash",
    "dietaryfiber",
    "carotene",
    "vitamin_a",
    "vitamin_e",
    "thiamin",
    "riboflavin",
    "niacin",
    "vitamin_c",
    "ca",
    "p",
    "k",
    "na",
    "mg",
    "fe",
    "zn",
    "se",
    "cu",
    "mn",
    "i",
    "sfa",
    "mufa",
    "pufa",
]


def _extract_nutrition_from_cns_row(row: Dict[str, Any]) -> Dict[str, Optional[float]]:
    def num(v: Any) -> Optional[float]:
        return _safe_num(v)

    return {
        "calorie": num(row.get("能量(kcal/100g)") or row.get("能量（kcal/100g）")),
        "protein": num(row.get("蛋白质(Protein)/g") or row.get("蛋白质（Protein）/g")),
        "fat": num(row.get("脂肪(Fat)/g") or row.get("脂肪（Fat）/g")),
        "carbohydrate": num(row.get("碳水化合物(CHO)/g") or row.get("碳水化合物（CHO）/g")),
        "water": num(row.get("水分(Water)/g") or row.get("水分（Water）/g")),
        "cholesterol": num(row.get("胆固醇(Cholesterol)/mg") or row.get("胆固醇（Cholesterol）/mg")),
        "ash": num(row.get("矿物质/g") or row.get("灰分(Ash)/g") or row.get("灰分（Ash）/g")),
        "dietaryfiber": num(row.get("总膳食纤维(Dietary fiber)/g") or row.get("总膳食纤维（Dietary fiber）/g")),
        "carotene": num(row.get("胡萝卜素(Carotene)/μg") or row.get("胡萝卜素（Carotene）/μg")),
        "vitamin_a": num(row.get("维生素A(Vitamin A)/μg") or row.get("维生素A（Vitamin A）/μg")),
        "vitamin_e": num(row.get("α-TE/mg")),
        "thiamin": num(row.get("硫胺素(Thiamin)/mg") or row.get("硫胺素（Thiamin）/mg")),
        "riboflavin": num(row.get("核黄素(Riboflavin)/mg") or row.get("核黄素（Riboflavin）/mg")),
        "niacin": num(row.get("烟酸(Niacin)/mg") or row.get("烟酸（Niacin）/mg")),
        "vitamin_c": num(row.get("维生素C(Vitamin C)/mg") or row.get("维生素C（Vitamin C）/mg")),
        "ca": num(row.get("钙(Ca)/mg") or row.get("钙（Ca）/mg")),
        "p": num(row.get("磷(P)/mg") or row.get("磷（P）/mg")),
        "k": num(row.get("钾(K)/mg") or row.get("钾（K）/mg")),
        "na": num(row.get("钠(Na)/mg") or row.get("钠（Na）/mg")),
        "mg": num(row.get("镁(Mg)/mg") or row.get("镁（Mg）/mg")),
        "fe": num(row.get("铁(Fe)/mg") or row.get("铁（Fe）/mg")),
        "zn": num(row.get("锌(Zn)/mg") or row.get("锌（Zn）/mg")),
        "se": num(row.get("硒(Se)/μg") or row.get("硒（Se）/μg")),
        "cu": num(row.get("铜(Cu)/mg") or row.get("铜（Cu）/mg")),
        "mn": num(row.get("锰(Mn)/mg") or row.get("锰（Mn）/mg")),
        "i": num(row.get("碘(I)/μg") or row.get("碘（I）/μg")),
        "sfa": num(row.get("饱和脂肪酸(SFA)/%") or row.get("饱和脂肪酸（SFA）/%")),
        "mufa": num(row.get("单不饱和脂肪酸(MUFA)/%") or row.get("单不饱和脂肪酸（MUFA）/%")),
        "pufa": num(row.get("多不饱和脂肪酸(PUFA)/%") or row.get("多不饱和脂肪酸（PUFA）/%")),
    }


@lru_cache(maxsize=2)
def _get_cns_rows_cached(workbook_path: str) -> List[Dict[str, Any]]:
    data = query_sheet_rows(workbook_path, "中国营养学会参考数据", q=None, mode="and", offset=0, limit=10**9)
    return data.get("rows") or []


@app.post("/api/nutrition/match_or_infer")
def nutrition_match_or_infer(payload: Dict[str, Any] = Body(...)):
    """
    nutrition 补全：先 CNS 匹配，失败再模型推断（返回标准 nutrition key）。
    """
    try:
        food_name = _norm_text(payload.get("food_name"))
        all_names = payload.get("all_names") or []
        names = [food_name] + [str(x).strip() for x in all_names if str(x).strip()]
        names = [_norm_text(x) for x in names if _norm_text(x)]
        if not names:
            raise ValueError("food_name/all_names 不能为空")

        p = get_workbook_path()
        cns_rows = _get_cns_rows_cached(p)
        idx_exact: Dict[str, Dict[str, Any]] = {}
        for r in cns_rows:
            nm = _norm_text(r.get("食物名称"))
            if nm and nm not in idx_exact:
                idx_exact[nm] = r

        picked = None
        picked_name = ""
        for n in names:
            if n in idx_exact:
                picked = idx_exact[n]
                picked_name = n
                break
        if picked is None:
            # 兜底：模糊匹配（较保守）
            best = (0.0, None, "")
            for n in names:
                for r in cns_rows:
                    rn = _norm_text(r.get("食物名称"))
                    if not rn:
                        continue
                    score = SequenceMatcher(None, n.lower(), rn.lower()).ratio()
                    if score > best[0]:
                        best = (score, r, rn)
            if best[1] is not None and best[0] >= 0.86:
                picked = best[1]
                picked_name = best[2]
        if picked is None:
            # 再兜底：包含匹配（例如“米饭”匹配“米饭（蒸，代表值）”）
            for n in names:
                nl = n.lower()
                for r in cns_rows:
                    rn = _norm_text(r.get("食物名称"))
                    if not rn:
                        continue
                    rnl = rn.lower()
                    if nl in rnl or rnl in nl:
                        picked = r
                        picked_name = rn
                        break
                if picked is not None:
                    break

        if picked is not None:
            return {
                "source": "cns",
                "matched_name": picked_name,
                "row": picked,
                "nutrition": _extract_nutrition_from_cns_row(picked),
            }

        # 模型推断
        try:
            obj = _ark_request_json(
                {
                    "task": "infer nutrition per 100g for a food item",
                    "food_name": food_name,
                    "all_names": names,
                    "keys": NUT_KEYS,
                    "output_schema": {
                        "nutrition": {k: 0 for k in NUT_KEYS},
                        "confidence": 0.75,
                        "reason": "简短理由",
                    },
                },
                "你是营养数据库助手。请仅输出 JSON；nutrition 字段必须包含给定 keys，无法判断时可给 null。",
            )
        except Exception:
            return {
                "source": "none",
                "nutrition": {k: None for k in NUT_KEYS},
                "confidence": 0.0,
                "reason": "CNS未匹配且模型不可用",
            }
        n = obj.get("nutrition") if isinstance(obj, dict) else None
        if not isinstance(n, dict):
            raise RuntimeError("模型返回缺少 nutrition")
        out_n: Dict[str, Optional[float]] = {}
        for k in NUT_KEYS:
            out_n[k] = _safe_num(n.get(k))
        conf = _safe_num(obj.get("confidence"))
        return {
            "source": "model",
            "nutrition": out_n,
            "confidence": max(0.0, min(1.0, conf if conf is not None else 0.0)),
            "reason": _norm_text(obj.get("reason")),
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/workbook/backup")
def workbook_backup():
    """
    创建当前 workbook 的备份文件，用于“撤回/还原”。
    """
    try:
        p = Path(get_workbook_path())
        if not p.exists():
            raise FileNotFoundError(str(p))
        backup_dir = APP_DIR / ".backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_id = f"{p.stem}.{ts}.bak.xlsx"
        out = backup_dir / backup_id
        shutil.copy2(str(p), str(out))
        return {"backup_id": backup_id}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/workbook/restore")
def workbook_restore(payload: Dict[str, Any] = Body(...)):
    """
    从备份恢复 workbook（覆盖写回 WORKBOOK_PATH 指向的文件）。
    """
    try:
        backup_id = str(payload.get("backup_id") or "").strip()
        if not backup_id:
            raise ValueError("backup_id 不能为空")
        backup_dir = APP_DIR / ".backups"
        src = backup_dir / backup_id
        if not src.exists():
            raise FileNotFoundError(str(src))
        dst = Path(get_workbook_path())
        shutil.copy2(str(src), str(dst))
        clear_caches()
        return {"restored": True, "backup_id": backup_id}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


