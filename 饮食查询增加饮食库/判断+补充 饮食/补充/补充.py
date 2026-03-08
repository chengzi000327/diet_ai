#!/usr/bin/env python3
from __future__ import annotations

import argparse
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook


DEFAULT_SOURCE_XLSX = "/Users/chengzi/Documents/work_space/饮食查询增加饮食库/判断+补充 饮食/补充/现有单位.xlsx"
DEFAULT_TARGET_XLSX = "/Users/chengzi/Documents/work_space/饮食查询增加饮食库/判断+补充 饮食/补充/需要的数据0228.xlsx"


def norm_text(v: object) -> str:
    if v is None:
        return ""
    return str(v).strip().lower().replace("（", "(").replace("）", ")")


def find_col(headers: list[object], names: list[str]) -> int:
    normalized = [str(h).strip() if h is not None else "" for h in headers]
    for n in names:
        if n in normalized:
            return normalized.index(n) + 1
    raise ValueError(f"未找到列: {names}；实际列: {normalized}")


def build_source_index(source_path: Path) -> Tuple[Dict[str, str], Dict[str, List[Tuple[str, object, object]]]]:
    """
    返回:
    - food_alias_map: food_name -> 别名
    - food_units_map: food_name -> [(unit, 单位转换（100克）, 单位别名), ...]
    """
    wb = load_workbook(source_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in ws[1]]

    food_col = find_col(headers, ["food_name", "食物名称"])
    alias_col = find_col(headers, ["别名"])
    unit_col = find_col(headers, ["unit", "单位"])
    unit_alias_col = find_col(headers, ["单位别名"])
    coef_col = find_col(headers, ["单位转换（100克）", "单位转换（100g）"])

    food_alias_map: Dict[str, str] = {}
    food_units_map: Dict[str, List[Tuple[str, object, object]]] = {}
    seen_unit_keys: Dict[str, set[str]] = {}

    for r in range(2, ws.max_row + 1):
        food_name = ws.cell(r, food_col).value
        unit = ws.cell(r, unit_col).value
        if food_name is None or unit is None:
            continue

        n_food = norm_text(food_name)
        n_unit = norm_text(unit)
        if not n_food or not n_unit:
            continue

        alias_val = ws.cell(r, alias_col).value
        if alias_val not in (None, "") and n_food not in food_alias_map:
            food_alias_map[n_food] = str(alias_val).strip()

        coef_val = ws.cell(r, coef_col).value
        unit_alias_val = ws.cell(r, unit_alias_col).value
        if n_food not in food_units_map:
            food_units_map[n_food] = []
            seen_unit_keys[n_food] = set()

        # 同一食物下，unit 去重保序
        if n_unit in seen_unit_keys[n_food]:
            continue
        seen_unit_keys[n_food].add(n_unit)
        food_units_map[n_food].append((str(unit).strip(), coef_val, unit_alias_val))

    return food_alias_map, food_units_map


def fill_target(target_path: Path, food_alias_map: Dict[str, str], food_units_map: Dict[str, List[Tuple[str, object, object]]]) -> dict:
    wb = load_workbook(target_path)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in ws[2]]  # 第2行是表头

    food_col = find_col(headers, ["food_name", "食物名称"])
    exists_col = find_col(headers, ["是否存在营养素"])
    alias_col = find_col(headers, ["别名"])
    unit_col = find_col(headers, ["单位", "unit"])
    coef_col = find_col(headers, ["单位转换（100g）", "单位转换（100克）"])
    unit_alias_col = find_col(headers, ["单位别名"])

    max_col = ws.max_column

    # 1) 只读取“食物首行”（A列有值）
    base_food_rows: List[List[object]] = []
    for r in range(3, ws.max_row + 1):
        food_name = ws.cell(r, food_col).value
        if food_name in (None, ""):
            continue
        row_vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        base_food_rows.append(row_vals)

    # 2) 重建成“米饭那种分组样式”
    rebuilt_rows: List[List[object]] = []
    block_lengths: List[int] = []
    expanded_foods = 0
    for row_vals in base_food_rows:
        food_name_val = row_vals[food_col - 1]
        n_food = norm_text(food_name_val)
        units = food_units_map.get(n_food, [])
        if not units:
            rebuilt_rows.append(row_vals)
            block_lengths.append(1)
            continue

        expanded_foods += 1
        block_lengths.append(len(units))
        alias_val = food_alias_map.get(n_food)

        for i, (unit_name, coef_val, unit_alias_val) in enumerate(units):
            new_row = [None] * max_col
            if i == 0:
                # 首行保留原有列，再覆盖别名/单位信息
                new_row = list(row_vals)
                if new_row[alias_col - 1] in (None, "") and alias_val:
                    new_row[alias_col - 1] = alias_val
            else:
                # 非首行按米饭格式，仅保留单位相关列
                new_row[food_col - 1] = None
                new_row[exists_col - 1] = None
                new_row[alias_col - 1] = None

            new_row[unit_col - 1] = unit_name
            new_row[coef_col - 1] = coef_val
            new_row[unit_alias_col - 1] = unit_alias_val
            rebuilt_rows.append(new_row)

    # 3) 清空旧数据区、取消旧合并，写入新数据
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    ws.delete_rows(3, ws.max_row - 2)

    for row_vals in rebuilt_rows:
        ws.append(row_vals)

    # 4) 对每个食物块做纵向合并（按米饭样式的关键列）
    # 仅合并 food_name/是否存在营养素/别名 三列，避免大规模合并导致性能过慢。
    merge_cols = [food_col, exists_col, alias_col]
    cursor = 3
    merged_blocks = 0
    for block_len in block_lengths:
        if block_len > 1:
            start_row = cursor
            end_row = cursor + block_len - 1
            for c in merge_cols:
                ws.merge_cells(start_row=start_row, start_column=c, end_row=end_row, end_column=c)
            merged_blocks += 1
        cursor += block_len

    wb.save(target_path)
    return {
        "foods_total": len(base_food_rows),
        "foods_expanded": expanded_foods,
        "rows_after_rebuild": len(rebuilt_rows),
        "merged_blocks": merged_blocks,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="将现有单位表中的单位信息补充到目标表。")
    parser.add_argument("--source", default=DEFAULT_SOURCE_XLSX, help="现有单位 Excel")
    parser.add_argument("--target", default=DEFAULT_TARGET_XLSX, help="需要补充的 Excel")
    args = parser.parse_args()

    source_path = Path(args.source)
    target_path = Path(args.target)

    food_alias_map, food_units_map = build_source_index(source_path)
    stats = fill_target(target_path, food_alias_map, food_units_map)

    print(f"完成: {target_path}")
    print(f"食物索引数: {len(food_alias_map)}")
    print(f"食物单位组索引数: {sum(len(v) for v in food_units_map.values())}")
    for k, v in stats.items():
        print(f"{k}: {v}")


if __name__ == "__main__":
    main()
