#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook


DEFAULT_QUERY_XLSX = "/Users/chengzi/Documents/work_space/饮食查询增加饮食库/判断缺失饮食/判断.xlsx"
DEFAULT_EXISTS_XLSX = "/Users/chengzi/Documents/work_space/饮食查询增加饮食库/判断缺失饮食/存在的食物.xlsx"

IDENTITY_MODIFIERS = [
    "去皮",
    "带皮",
    "糯",
    "团",
    "脱脂",
    "低脂",
    "全脂",
    "无糖",
    "加糖",
    "蒸",
    "煮",
    "炸",
    "烤",
    "卤",
    "炒",
    "拌",
    "生",
    "熟",
    "草莓味",
    "原味",
    "香草味",
    "巧克力味",
]

SYNONYM_REPLACEMENTS = [
    ("橙子", "橙")
]


def normalize_text(text: str) -> str:
    s = (text or "").strip().lower()
    s = s.replace("（", "(").replace("）", ")")
    s = s.replace("【", "[").replace("】", "]")
    s = re.sub(r"\s+", "", s)
    return s


def build_char_ngrams(text: str) -> List[str]:
    # 同时使用 1/2/3-gram，兼顾短词（如“米饭”）和长词（如“小麦粉(标准粉)”）
    grams: List[str] = []
    chars = list(text)
    for n in (1, 2, 3):
        if len(chars) < n:
            continue
        for i in range(len(chars) - n + 1):
            grams.append("".join(chars[i : i + n]))
    return grams


def strip_parenthesized(text: str) -> str:
    s = normalize_text(text)
    # 去掉括号说明词，如：米饭(蒸)(均值) -> 米饭
    s = re.sub(r"\([^)]*\)", "", s)
    return s


def apply_synonym_replacements(text: str) -> str:
    s = text
    for src, dst in SYNONYM_REPLACEMENTS:
        s = s.replace(src, dst)
    return s


def canonical_name(text: str) -> str:
    # 用于 embedding 的规范名：去括号说明 + 同义词替换（保留加工/口味词）
    s = strip_parenthesized(text)
    s = apply_synonym_replacements(s)
    return s


def remove_modifier_tokens(text: str) -> str:
    s = canonical_name(text)
    for token in IDENTITY_MODIFIERS:
        s = s.replace(token, "")
    return s


def identity_modifier_set(text: str) -> set[str]:
    # 保留括号内内容，像“酸奶(脱脂)”中的“脱脂”需要被识别出来
    s = normalize_text(text)
    return {token for token in IDENTITY_MODIFIERS if token in s}


def token_hash(token: str, dim: int) -> int:
    digest = hashlib.md5(token.encode("utf-8")).hexdigest()
    return int(digest, 16) % dim


def to_sparse_embedding(text: str, dim: int = 1024) -> Dict[int, float]:
    normalized = canonical_name(text)
    grams = build_char_ngrams(normalized)
    vec: Dict[int, float] = {}
    for gram in grams:
        idx = token_hash(gram, dim)
        vec[idx] = vec.get(idx, 0.0) + 1.0
    return vec


def cosine_sparse(v1: Dict[int, float], v2: Dict[int, float]) -> float:
    if not v1 or not v2:
        return 0.0

    dot = 0.0
    # iterate smaller dict for speed
    if len(v1) > len(v2):
        v1, v2 = v2, v1
    for k, x in v1.items():
        y = v2.get(k)
        if y is not None:
            dot += x * y

    n1 = math.sqrt(sum(x * x for x in v1.values()))
    n2 = math.sqrt(sum(x * x for x in v2.values()))
    if n1 == 0.0 or n2 == 0.0:
        return 0.0
    return dot / (n1 * n2)


@dataclass
class MatchResult:
    exists: bool
    best_name: str
    score: float


@dataclass
class Candidate:
    name: str
    vec: Dict[int, float]
    nutrition: Optional[Dict[str, float]]


def exact_match(a: str, b: str) -> bool:
    # 只把“完全相同”视为满分命中，避免“鸡腿 vs 去皮鸡腿”被直接判 1 分
    return bool(a and b and a == b)


def safe_ratio_diff(a: float, b: float) -> float:
    base = max(abs(a), abs(b), 1e-9)
    return abs(a - b) / base


def nutrition_similarity(q: Optional[Dict[str, float]], c: Optional[Dict[str, float]]) -> Optional[float]:
    if not q or not c:
        return None
    keys = ("calorie", "protein", "fat", "carbohydrate")
    pairs = []
    for k in keys:
        if k in q and k in c:
            pairs.append((q[k], c[k]))
    if not pairs:
        return None

    # 差异越小越接近 1，差异很大时接近 0
    diffs = [safe_ratio_diff(a, b) for a, b in pairs]
    avg = sum(diffs) / len(diffs)
    return max(0.0, 1.0 - min(avg, 1.0))


def best_match(
    query: str,
    query_nutrition: Optional[Dict[str, float]],
    candidates: List[Candidate],
    threshold: float,
) -> MatchResult:
    nq = normalize_text(query)
    nq_core = canonical_name(query)
    qv = to_sparse_embedding(query)

    best_name = ""
    best_score = -1.0
    for cand in candidates:
        nc = normalize_text(cand.name)
        nc_core = canonical_name(cand.name)
        if exact_match(nq, nc):
            return MatchResult(True, cand.name, 1.0)
        # xx 与 xx(均值) 视为同名，直接 1 分
        if nq_core == nc_core and ("均值" in nq or "均值" in nc):
            return MatchResult(True, cand.name, 1.0)
        # 主词一致（如 橙子 vs 橙）给高分；
        # 但如果存在加工/口味/身份修饰词差异（如 酸奶 vs 酸奶(脱脂)、卤蛋 vs 蛋），显著降分。
        if nq_core and nq_core == nc_core:
            qmods = identity_modifier_set(nq)
            cmods = identity_modifier_set(nc)
            diff_count = len((qmods - cmods) | (cmods - qmods))
            if diff_count > 0:
                # 有修饰词差异时，不允许进入“有”
                return MatchResult(False, cand.name, max(0.0, 0.85 - 0.12 * (diff_count - 1)))
            return MatchResult(True, cand.name, 0.995)

        name_score = cosine_sparse(qv, cand.vec)
        # 对“包含关系”施加惩罚，避免过高分
        if nq in nc or nc in nq:
            name_score *= 0.88

        # 语义主词一致性约束：
        # - 去掉括号和修饰词后，如果主词都非空且不一致，降低得分
        # - 完全一致时加一点奖励
        if nq_core and nc_core:
            if nq_core == nc_core:
                name_score *= 1.2
            elif nq_core not in nc_core and nc_core not in nq_core:
                name_score *= 0.72

        # 身份修饰词差异惩罚：例如 鸡腿 vs 去皮鸡腿，米饭 vs 糯米饭团
        qmods = identity_modifier_set(nq)
        cmods = identity_modifier_set(nc)
        diff_count = len((qmods - cmods) | (cmods - qmods))
        if diff_count > 0:
            name_score *= 0.68**diff_count

        n_score = nutrition_similarity(query_nutrition, cand.nutrition)
        score = name_score if n_score is None else (0.8 * name_score + 0.2 * n_score)
        if score > best_score:
            best_score = score
            best_name = cand.name

    final_score = best_score if best_score > 0 else 0.0
    return MatchResult(final_score > threshold, best_name, final_score)


def find_column_idx(header_row: List[object], candidates: List[str]) -> int:
    normalized = [str(x).strip() if x is not None else "" for x in header_row]
    for name in candidates:
        if name in normalized:
            return normalized.index(name) + 1  # openpyxl is 1-indexed
    raise ValueError(f"未找到列名，候选: {candidates}，实际: {normalized}")


def find_optional_column_idx(header_row: List[object], candidates: List[str]) -> Optional[int]:
    normalized = [str(x).strip() if x is not None else "" for x in header_row]
    for name in candidates:
        if name in normalized:
            return normalized.index(name) + 1
    return None


def read_nutrition_row(ws, row: int, col_map: Dict[str, Optional[int]]) -> Optional[Dict[str, float]]:
    out: Dict[str, float] = {}
    for k, col in col_map.items():
        if col is None:
            continue
        v = ws.cell(row, col).value
        if v is None or str(v).strip() == "":
            continue
        try:
            out[k] = float(v)
        except Exception:
            continue
    return out or None


def main() -> None:
    parser = argparse.ArgumentParser(description="使用 embedding 判断食物是否存在，并回填“是否存在营养素”列。")
    parser.add_argument("--query-xlsx", default=DEFAULT_QUERY_XLSX, help="待判断文件（包含 food_name）")
    parser.add_argument("--exists-xlsx", default=DEFAULT_EXISTS_XLSX, help="存在食物文件（包含 食物名称）")
    parser.add_argument("--threshold", type=float, default=0.98, help="embedding 相似度阈值（严格大于），默认 0.98")
    args = parser.parse_args()

    query_path = Path(args.query_xlsx)
    exists_path = Path(args.exists_xlsx)

    wb_exists = load_workbook(exists_path, data_only=True)
    ws_exists = wb_exists[wb_exists.sheetnames[0]]

    # 读取“存在的食物”
    exists_header = [cell.value for cell in ws_exists[1]]
    exists_name_col = find_column_idx(exists_header, ["食物名称", "food_name", "name"])
    exists_nut_cols = {
        "calorie": find_optional_column_idx(exists_header, ["calorie_per_100g", "卡路里", "能量(kcal/100g)", "calorie"]),
        "protein": find_optional_column_idx(exists_header, ["protein_per_100g", "蛋白质", "蛋白质(Protein)/g", "protein"]),
        "fat": find_optional_column_idx(exists_header, ["fat_per_100g", "脂肪", "脂肪(Fat)/g", "fat"]),
        "carbohydrate": find_optional_column_idx(
            exists_header,
            ["carbohydrate_per_100g", "碳水化合物", "碳水化合物(CHO)/g", "carbohydrate"],
        ),
    }

    exists_records: List[tuple[str, Optional[Dict[str, float]]]] = []
    for r in range(2, ws_exists.max_row + 1):
        val = ws_exists.cell(r, exists_name_col).value
        if val is None:
            continue
        s = str(val).strip()
        if not s:
            continue
        exists_records.append((s, read_nutrition_row(ws_exists, r, exists_nut_cols)))

    # 去重保序
    deduped_exists: List[tuple[str, Optional[Dict[str, float]]]] = []
    seen = set()
    for name, nut in exists_records:
        if name not in seen:
            deduped_exists.append((name, nut))
            seen.add(name)

    candidate_embeddings = [Candidate(name=n, vec=to_sparse_embedding(n), nutrition=nut) for n, nut in deduped_exists]

    wb_query = load_workbook(query_path)
    ws_query = wb_query[wb_query.sheetnames[0]]

    query_header = [cell.value for cell in ws_query[1]]
    food_col = find_column_idx(query_header, ["food_name", "食物名称", "name"])
    query_nut_cols = {
        "calorie": find_optional_column_idx(query_header, ["calorie_per_100g", "卡路里", "能量(kcal/100g)", "calorie"]),
        "protein": find_optional_column_idx(query_header, ["protein_per_100g", "蛋白质", "蛋白质(Protein)/g", "protein"]),
        "fat": find_optional_column_idx(query_header, ["fat_per_100g", "脂肪", "脂肪(Fat)/g", "fat"]),
        "carbohydrate": find_optional_column_idx(
            query_header,
            ["carbohydrate_per_100g", "碳水化合物", "碳水化合物(CHO)/g", "carbohydrate"],
        ),
    }

    # 目标列：是否存在营养素（没有就新增）
    try:
        exists_col = find_column_idx(query_header, ["是否存在营养素"])
    except ValueError:
        exists_col = ws_query.max_column + 1
        ws_query.cell(1, exists_col).value = "是否存在营养素"

    # 额外输出辅助列，方便人工核查
    helper_match_col = ws_query.max_column + 1
    ws_query.cell(1, helper_match_col).value = "匹配到的食物名称"
    helper_score_col = ws_query.max_column + 1
    ws_query.cell(1, helper_score_col).value = "匹配相似度"

    total = 0
    yes_count = 0
    no_count = 0

    for r in range(2, ws_query.max_row + 1):
        food = ws_query.cell(r, food_col).value
        if food is None or not str(food).strip():
            continue
        total += 1
        q_nutrition = read_nutrition_row(ws_query, r, query_nut_cols)
        result = best_match(str(food), q_nutrition, candidate_embeddings, threshold=args.threshold)
        ws_query.cell(r, exists_col).value = "有" if result.exists else "无"
        ws_query.cell(r, helper_match_col).value = result.best_name
        ws_query.cell(r, helper_score_col).value = round(result.score, 4)
        if result.exists:
            yes_count += 1
        else:
            no_count += 1

    wb_query.save(query_path)

    print(f"完成: {query_path}")
    print(f"阈值: {args.threshold}")
    print(f"总计判断: {total}")
    print(f"有: {yes_count}")
    print(f"无: {no_count}")


if __name__ == "__main__":
    main()
