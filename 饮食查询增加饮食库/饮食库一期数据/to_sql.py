#!/usr/bin/env python3
"""
从 Excel 中提取 (food_name, unit) 对，并展开“别名”。

示例：
输入行:
  food_name = 米饭
  别名 = 白米饭，大米饭
  unit = 碗
输出对:
  (米饭, 碗), (白米饭, 碗), (大米饭, 碗)
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Iterable, List, Sequence, Tuple

from openpyxl import load_workbook

Pair = Tuple[str, str]
DELIMITERS = {"，", ",", "、", "/", ";", "；", "|"}
OPEN_TO_CLOSE = {"(": ")", "（": "）", "[": "]", "【": "】", "{": "}"}
CLOSE_SET = set(OPEN_TO_CLOSE.values())


def split_multi_value(text: str | None) -> List[str]:
    if text is None:
        return []
    raw = str(text).strip()
    if not raw:
        return []

    parts: List[str] = []
    token: List[str] = []
    stack: List[str] = []

    for ch in raw:
        if ch in OPEN_TO_CLOSE:
            stack.append(OPEN_TO_CLOSE[ch])
            token.append(ch)
            continue

        if ch in CLOSE_SET:
            if stack and ch == stack[-1]:
                stack.pop()
            token.append(ch)
            continue

        if ch in DELIMITERS and not stack:
            item = "".join(token).strip()
            if item:
                parts.append(item)
            token = []
            continue

        token.append(ch)

    tail = "".join(token).strip()
    if tail:
        parts.append(tail)

    return parts


def unique_keep_order(items: Iterable[str]) -> List[str]:
    seen = set()
    result: List[str] = []
    for item in items:
        if item not in seen:
            seen.add(item)
            result.append(item)
    return result


def extract_pairs(excel_path: Path, sheet_name: str | None = None) -> List[Pair]:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    headers = next(rows)
    if headers is None:
        raise ValueError("Excel 表头为空。")

    header_to_index = {str(h).strip(): idx for idx, h in enumerate(headers) if h is not None}
    required = ["food_name", "别名", "unit"]
    missing = [name for name in required if name not in header_to_index]
    if missing:
        raise ValueError(f"缺少必要列: {missing}，实际列: {list(header_to_index.keys())}")

    food_idx = header_to_index["food_name"]
    alias_idx = header_to_index["别名"]
    unit_idx = header_to_index["unit"]

    pairs: List[Pair] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue

        food_name = str(row[food_idx]).strip() if row[food_idx] is not None else ""
        alias_text = str(row[alias_idx]).strip() if row[alias_idx] is not None else ""
        unit_text = str(row[unit_idx]).strip() if row[unit_idx] is not None else ""

        if not unit_text:
            continue

        names = unique_keep_order([food_name] + split_multi_value(alias_text))
        units = unique_keep_order(split_multi_value(unit_text) or [unit_text])

        for name in names:
            if not name:
                continue
            for unit in units:
                if not unit:
                    continue
                pairs.append((name, unit))

    return unique_keep_order(pairs)


def to_sql_insert(pairs: Sequence[Pair], table: str) -> str:
    values = []
    for food_name, unit in pairs:
        safe_food_name = food_name.replace("'", "''")
        safe_unit = unit.replace("'", "''")
        values.append(f"('{safe_food_name}', '{safe_unit}')")
    joined = ",\n".join(values)
    return f"INSERT INTO {table} (food_name, unit) VALUES\n{joined};\n"


def write_csv(pairs: Sequence[Pair], output_path: Path) -> None:
    with output_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["food_name", "unit"])
        writer.writerows(pairs)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="展开 Excel 中 food_name/别名 与 unit 的一一对应关系。"
    )
    parser.add_argument("excel_path", type=Path, help="Excel 文件路径")
    parser.add_argument("--sheet", type=str, default=None, help="Sheet 名称（默认第一个）")
    parser.add_argument(
        "--out-csv",
        type=Path,
        default=Path("food_unit_pairs.csv"),
        help="输出 CSV 路径（默认 food_unit_pairs.csv）",
    )
    parser.add_argument(
        "--out-sql",
        type=Path,
        default=Path("food_unit_pairs.sql"),
        help="输出 SQL 路径（默认 food_unit_pairs.sql）",
    )
    parser.add_argument(
        "--table",
        type=str,
        default="food_unit_mapping",
        help="SQL 目标表名（默认 food_unit_mapping）",
    )
    parser.add_argument(
        "--print-sample",
        type=int,
        default=20,
        help="终端打印前 N 条结果（默认 20）",
    )
    args = parser.parse_args()

    pairs = extract_pairs(args.excel_path, args.sheet)
    write_csv(pairs, args.out_csv)
    args.out_sql.write_text(to_sql_insert(pairs, args.table), encoding="utf-8")

    print(f"总共生成 {len(pairs)} 对 (food_name, unit)")
    print(f"CSV 已写出: {args.out_csv}")
    print(f"SQL 已写出: {args.out_sql}")
    print("")
    print("样例：")
    for food_name, unit in pairs[: max(0, args.print_sample)]:
        print(f"({food_name}, {unit})")


if __name__ == "__main__":
    main()
